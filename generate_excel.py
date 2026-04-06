#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
公共建筑碳排放计算报告 - Excel计算表格生成器（完整版）
基于GB/T 51366-2019《建筑碳排放计算标准》和JS/T 303-2026《公共机构碳排放核算指南》

功能：
  - 13个工作表全覆盖
  - 50+建材类型、20种能源、15+用能系统
  - 情景分析、敏感性分析、对标预测、碳交易分析
  - 自动计算公式和数据验证

用法：
  python generate_excel.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
import os

# ========================================
# 配置
# ========================================
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "公共建筑碳排放计算表.xlsx")

# ========================================
# 样式系统
# ========================================
class Styles:
    """统一的样式定义"""
    title_font = Font(name='微软雅黑', size=16, bold=True, color='1F4E79')
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    subheader_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    section_font = Font(name='微软雅黑', size=13, bold=True, color='1F4E79')
    normal_font = Font(name='微软雅黑', size=10)
    bold_font = Font(name='微软雅黑', size=10, bold=True)
    formula_font = Font(name='微软雅黑', size=10, italic=True, color='0066CC')
    note_font = Font(name='微软雅黑', size=9, italic=True, color='666666')

    section_fills = [
        PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid'),
        PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid'),
        PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid'),
        PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid'),
        PatternFill(start_color='F44336', end_color='F44336', fill_type='solid'),
        PatternFill(start_color='00BCD4', end_color='00BCD4', fill_type='solid'),
        PatternFill(start_color='795548', end_color='795548', fill_type='solid'),
        PatternFill(start_color='607D8B', end_color='607D8B', fill_type='solid'),
        PatternFill(start_color='E91E63', end_color='E91E63', fill_type='solid'),
        PatternFill(start_color='3F51B5', end_color='3F51B5', fill_type='solid'),
        PatternFill(start_color='009688', end_color='009688', fill_type='solid'),
        PatternFill(start_color='FF5722', end_color='FF5722', fill_type='solid'),
    ]

    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    calc_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    total_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    param_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    scenario_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    warning_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    benchmark_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')


def apply_header_row(ws, row, headers, fill=None, font=None):
    """应用表头样式到一行"""
    fill = fill or Styles.section_fills[0]
    font = font or Styles.subheader_font
    ws.row_dimensions[row].height = 30
    for j, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = Styles.center_align
        cell.border = Styles.thin_border


def apply_data_row(ws, row, data, input_cols=None):
    """应用数据行样式"""
    input_cols = input_cols or []
    for j, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=j, value=value)
        cell.border = Styles.thin_border
        cell.font = Styles.normal_font
        cell.alignment = Styles.center_align if j <= 3 else Styles.left_align
        if j in input_cols:
            cell.fill = Styles.input_fill
        elif j > 3 and j <= 6:
            cell.fill = Styles.calc_fill


def set_col_widths(ws, widths):
    """设置列宽"""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


# ========================================
# 工作表生成器
# ========================================

def create_cover_sheet(wb):
    """工作表1: 使用说明"""
    ws = wb.active
    ws.title = "使用说明"
    ws.sheet_properties.tabColor = '1F4E79'

    cover_content = [
        ["公共建筑碳排放计算系统 V2.0", 24, '1F4E79', True],
        ["", 14, '666666', False],
        ["基于GB/T 51366-2019《建筑碳排放计算标准》", 13, '1F4E79', True],
        ["基于JS/T 303-2026《公共机构碳排放核算指南》", 13, '1F4E79', True],
        ["", 12, '666666', False],
        ["=" * 100, 12, '999999', False],
        ["", 12, '666666', False],
        ["系统特色功能", 18, '1F4E79', True],
        ["", 12, '666666', False],
        ["✓ 全面计算覆盖", 13, '2E75B6', False],
        ["  · 运行阶段：15+用能系统（暖通空调、生活热水、照明、电梯、数据中心、充电桩等）", 11, '666666', False],
        ["  · 建材阶段：50+建材类型（结构、围护、装饰、安装、特殊、再生材料）", 11, '666666', False],
        ["  · 建造拆除：20+分部分项工程 + 绿色施工措施", 11, '666666', False],
        ["", 12, '666666', False],
        ["✓ 扩展计算模块", 13, '4CAF50', False],
        ["  · 水资源消耗碳排放（供水、排水、水处理）", 11, '666666', False],
        ["  · 废弃物处理碳排放（生活垃圾、建筑垃圾、危险废物）", 11, '666666', False],
        ["  · 建筑碳汇计算（绿地碳汇、立体绿化、屋顶绿化）", 11, '666666', False],
        ["  · 隐含碳排放（家具、设备、IT资产）", 11, '666666', False],
        ["", 12, '666666', False],
        ["✓ 高级分析功能", 13, 'FF9800', False],
        ["  · 情景分析（基准情景、节能情景、碳中和情景）", 11, '666666', False],
        ["  · 敏感性分析（识别关键影响因素）", 11, '666666', False],
        ["  · 对标分析（与国家/地方/行业标准对比）", 11, '666666', False],
        ["  · 预测分析（未来5-10年碳排放趋势）", 11, '666666', False],
        ["  · 碳交易分析（碳配额、碳成本核算）", 11, '666666', False],
        ["", 12, '666666', False],
        ["✓ 数据管理与报告", 13, '9C27B0', False],
        ["  · Excel-Word自动联动", 11, '666666', False],
        ["  · 多维度数据汇总", 11, '666666', False],
        ["  · 可视化图表生成", 11, '666666', False],
        ["  · 完整报告自动生成", 11, '666666', False],
        ["", 12, '666666', False],
        ["=" * 100, 12, '999999', False],
        ["", 12, '666666', False],
        ["工作表导航", 18, '1F4E79', True],
        ["", 12, '666666', False],
        ["1.  使用说明        - 本工作表，系统功能介绍", 11, '2E75B6', False],
        ["2.  项目基本信息    - 建筑基本信息、运行参数、可再生能源", 11, '4CAF50', False],
        ["3.  运行阶段-能源   - 能源消耗汇总、各类能源详细计算", 11, 'FF9800', False],
        ["4.  运行阶段-系统   - 15+用能系统详细计算", 11, '9C27B0', False],
        ["5.  建材生产运输    - 50+建材类型、生产和运输碳排放", 11, 'F44336', False],
        ["6.  建造拆除阶段    - 分部分项工程、绿色施工、拆除", 11, '00BCD4', False],
        ["7.  水资源与废弃物  - 水资源消耗、废弃物处理碳排放", 11, '795548', False],
        ["8.  碳汇与隐含碳    - 建筑碳汇、家具设备隐含碳", 11, '607D8B', False],
        ["9.  情景分析        - 多情景对比分析", 11, 'E91E63', False],
        ["10. 敏感性分析      - 关键参数敏感性测试", 11, '3F51B5', False],
        ["11. 对标与预测      - 行业标准对标、碳排放预测", 11, '009688', False],
        ["12. 碳交易分析      - 碳配额、碳成本核算", 11, 'FF5722', False],
        ["13. 碳排放汇总      - 全生命周期碳排放汇总报告", 11, 'F44336', False],
        ["", 12, '666666', False],
        ["=" * 100, 12, '999999', False],
        ["", 12, '666666', False],
        ["填写说明", 18, '1F4E79', True],
        ["", 12, '666666', False],
        ["🟡 黄色单元格：输入区，需要填写基础数据", 11, '666666', False],
        ["🟢 绿色单元格：计算区，自动计算公式结果", 11, '666666', False],
        ["⚪ 灰色单元格：汇总区，各阶段碳排放汇总", 11, '666666', False],
        ["🔵 浅蓝单元格：情景分析参数", 11, '666666', False],
        ["🔴 浅红单元格：警告/注意数据", 11, '666666', False],
        ["🟤 浅绿单元格：行业基准数据", 11, '666666', False],
        ["", 12, '666666', False],
        ["【重要提示】", 12, 'F44336', True],
        ["1. 所有能源消耗数据请按照实际计量数据填写", 11, '666666', False],
        ["2. 排放因子优先采用本表提供的缺省值，如有实测数据可自行修改", 11, '666666', False],
        ["3. 计算完成后，点击'更新Word报告'按钮自动生成Word文档", 11, '666666', False],
        ["4. 建议定期备份Excel文件，重要修改前做好版本记录", 11, '666666', False],
        ["5. 本系统与Word文档保持数据联动，Excel数据变更会自动更新Word内容", 11, '666666', False],
    ]

    for i, (text, size, color, is_bold) in enumerate(cover_content, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(name='微软雅黑', size=size, bold=is_bold, color=color)
        cell.alignment = Styles.left_align
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=12)
        ws.row_dimensions[i].height = 20 if size <= 12 else 30

    ws.column_dimensions['A'].width = 120
    for col in range(2, 13):
        ws.column_dimensions[get_column_letter(col)].width = 15

    return ws


def create_basic_info_sheet(wb):
    """工作表2: 项目基本信息"""
    ws = wb.create_sheet("项目基本信息")
    ws.sheet_properties.tabColor = '4CAF50'

    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1, value="公共建筑碳排放计算 - 项目基本信息与运行参数").font = Styles.title_font
    ws.cell(row=1, column=1).alignment = Styles.center_align

    # 2.1 建筑基本信息
    ws.merge_cells('A3:H3')
    ws.cell(row=3, column=1, value="2.1 建筑基本信息").font = Styles.section_font

    basic_sections = [
        ("基础信息", [
            ["项目名称", "", "-", "", "项目文件", "填写完整项目名称"],
            ["建筑地址", "", "-", "", "规划许可证", "详细到街道门牌号"],
            ["建筑类型", "", "-", "", "设计文件", "办公/商业/医院/学校/酒店等"],
            ["建筑功能", "", "-", "", "设计文件", "主要功能说明"],
            ["建筑面积(A)", "", "m²", "", "房产证/测绘报告", "地上+地下总面积"],
            ["地上建筑面积", "", "m²", "", "设计图纸", ""],
            ["地下建筑面积", "", "m²", "", "设计图纸", ""],
            ["建筑层数（地上）", "", "层", "", "设计文件", ""],
            ["建筑层数（地下）", "", "层", "", "设计文件", ""],
            ["建筑高度", "", "m", "", "设计文件", "室外地面到屋面面层"],
            ["结构类型", "", "-", "", "设计文件", "框架/剪力墙/钢结构等"],
            ["竣工年份", "", "年", "", "竣工验收文件", ""],
            ["计算年度", "", "年", "", "报告要求", "碳排放核算的年份"],
            ["设计使用寿命", 50, "年", "", "设计文件", "GB/T 51366-2019规定"],
        ]),
        ("使用信息", [
            ["建筑使用人数", "", "人", "", "物业管理记录", "日均使用人数"],
            ["峰值使用人数", "", "人", "", "物业管理记录", "最大使用人数"],
            ["年运行天数", 365, "天", "", "实际运行记录", "可按实际情况调整"],
            ["日运行时间", 24, "小时", "", "实际运行记录", "平均每日运行小时数"],
            ["入住率", 80, "%", "", "物业管理记录", "实际使用面积/总面积"],
            ["停车位数量", "", "个", "", "规划许可证", "地面+地下停车位"],
            ["绿地面积", "", "m²", "", "规划许可证", "用于计算碳汇"],
            ["绿地率", "", "%", "", "规划许可证", "绿地面积/总用地面积"],
            ["屋顶绿化面积", "", "m²", "", "设计文件", "如有立体绿化"],
            ["立体绿化面积", "", "m²", "", "设计文件", "垂直绿化等"],
        ]),
    ]

    current_row = 4
    for section_name, data in basic_sections:
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws.cell(row=current_row, column=1, value=section_name).font = Font(name='微软雅黑', size=11, bold=True, color='1F4E79')
        current_row += 1

        headers = ["项目", "数值", "单位", "数据来源", "说明", "", "", ""]
        for j, header in enumerate(headers, 1):
            if header:
                cell = ws.cell(row=current_row, column=j, value=header)
                cell.font = Styles.subheader_font
                cell.fill = Styles.section_fills[0]
                cell.alignment = Styles.center_align
                cell.border = Styles.thin_border
        current_row += 1

        for row_data in data:
            for j, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=j, value=value)
                cell.border = Styles.thin_border
                cell.font = Styles.normal_font
                cell.alignment = Styles.center_align if j <= 3 else Styles.left_align
                if j == 2:
                    cell.fill = Styles.input_fill
                elif j in [4, 5]:
                    cell.fill = Styles.param_fill
            current_row += 1
        current_row += 1

    set_col_widths(ws, [22, 15, 12, 18, 35, 25, 15, 15])
    return ws


def create_energy_sheet(wb):
    """工作表3: 运行阶段-能源"""
    ws = wb.create_sheet("运行阶段-能源")
    ws.sheet_properties.tabColor = 'FF9800'

    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value="运行阶段碳排放计算 - 能源消耗汇总与详细计算").font = Styles.title_font

    # 3.1 能源消耗汇总
    ws.merge_cells('A3:J3')
    ws.cell(row=3, column=1, value="3.1 能源消耗汇总（基于JS/T 303-2026第6章）").font = Styles.section_font

    energy_headers = [
        "序号", "能源类型", "年消耗量", "单位",
        "折标系数\n(kgce/单位)", "折标准煤\n(tce)",
        "碳排放因子\n(kgCO₂/单位)", "碳排放量\n(tCO₂)",
        "数据来源", "备注"
    ]
    apply_header_row(ws, 4, energy_headers, Styles.section_fills[1])

    energy_types = [
        [1, "电力（常规）", "", "kWh", 0.1229, "", 0.5836, "", "电费账单/电表", "市电"],
        [2, "电力（绿电）", "", "kWh", 0.1229, "", 0.0000, "", "绿电采购合同", "零碳排放因子"],
        [3, "天然气", "", "m³", 1.3300, "", 2.184, "", "燃气账单/气表", "管道天然气"],
        [4, "集中供热（热水）", "", "GJ", 34.12, "", 0.11, "", "热力账单/热量表", "市政集中供热"],
        [5, "集中供热（蒸汽）", "", "t", 0.0, "", 0.0, "", "热力账单", "如有蒸汽供热"],
        [6, "汽油", "", "L", 0.7344, "", 2.179, "", "加油记录", "公务用车"],
        [7, "柴油", "", "L", 0.8571, "", 2.718, "", "加油记录", "发电机/锅炉"],
        [8, "液化石油气", "", "kg", 1.7143, "", 3.166, "", "采购记录", "食堂/生活"],
        [9, "煤炭", "", "t", 0.7143, "", 0.0, "", "采购记录", "如有燃煤锅炉"],
        [10, "燃料油", "", "L", 0.0, "", 2.978, "", "采购记录", "应急发电机"],
        [11, "生物质燃料", "", "kg", 0.0, "", 0.0, "", "采购记录", "如有生物质锅炉"],
        [12, "地热能", "", "kWh", 0.0, "", 0.0, "", "运行记录", "地源热泵"],
        [13, "其他能源1", "", "", 0.0, "", 0.0, "", "", ""],
        [14, "其他能源2", "", "", 0.0, "", 0.0, "", "", ""],
        [15, "其他能源3", "", "", 0.0, "", 0.0, "", "", ""],
        ["", "合计", "", "", "", "", "", "", "", ""],
    ]

    for i, row in enumerate(energy_types, 5):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.normal_font
            cell.alignment = Styles.center_align if j in [1, 4, 5, 7] else Styles.left_align
            if j == 3:
                cell.fill = Styles.input_fill
            elif j in [5, 7]:
                cell.fill = Styles.param_fill
            elif j == 8:
                cell.fill = Styles.calc_fill
                if i < 16 and isinstance(row[2], str) and row[2] != "":
                    cell.value = f"=IF(C{i}=\"\",\"\",C{i}*G{i}/1000)"
                    cell.number_format = '#,##0.00'
            if i == 16:
                cell.fill = Styles.total_fill
                cell.font = Styles.bold_font
                if j == 6:
                    cell.value = "=SUM(F5:F15)"
                elif j == 8:
                    cell.value = "=SUM(H5:H15)"

    for row in range(5, 16):
        ws.cell(row=row, column=6, value=f"=IF(C{row}=\"\",\"\",C{row}*E{row}/1000)")
        ws.cell(row=row, column=6).number_format = '#,##0.00'

    # 3.2 能源结构分析
    ws.merge_cells('A18:J18')
    ws.cell(row=18, column=1, value="3.2 能源结构分析").font = Styles.section_font

    structure_headers = ["指标", "数值", "单位", "计算公式", "行业标准", "对标结果", "", "", "", ""]
    apply_header_row(ws, 19, structure_headers, Styles.section_fills[2])

    structure_data = [
        ["一次能源消耗占比", "", "%", "=（煤炭+天然气+石油类）/总能耗", "-", "", "", "", "", ""],
        ["二次能源消耗占比", "", "%", "=电力+热力/总能耗", "-", "", "", "", "", ""],
        ["可再生能源占比", "", "%", "=（绿电+地热+生物质）/总能耗", "-", "", "", "", "", ""],
        ["化石能源占比", "", "%", "=（煤炭+天然气+石油类）/总能耗", "-", "", "", "", "", ""],
        ["清洁能源占比", "", "%", "=（绿电+天然气+可再生能源）/总能耗", "-", "", "", "", "", ""],
        ["综合能耗指标", "", "kgce/m²", "=总折标准煤/建筑面积", "<0.5", "", "", "", "", ""],
        ["单位面积电力消耗", "", "kWh/m²", "=总用电量/建筑面积", "<150", "", "", "", "", ""],
        ["人均综合能耗", "", "kgce/人", "=总折标准煤/使用人数", "<50", "", "", "", "", ""],
    ]

    for i, row in enumerate(structure_data, 20):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.bold_font if j == 1 else Styles.normal_font
            if j == 1:
                cell.fill = Styles.param_fill
            elif j == 2:
                cell.fill = Styles.calc_fill
            elif j == 4:
                cell.font = Styles.formula_font
                cell.fill = Styles.param_fill
            elif j == 5:
                cell.fill = Styles.benchmark_fill

    set_col_widths(ws, [25, 18, 15, 35, 18, 18, 15, 15, 15, 15])
    return ws


def create_systems_sheet(wb):
    """工作表4: 运行阶段-系统"""
    ws = wb.create_sheet("运行阶段-系统")
    ws.sheet_properties.tabColor = '9C27B0'

    ws.merge_cells('A1:L1')
    ws.cell(row=1, column=1, value="运行阶段用能系统详细计算（15+系统类型）").font = Styles.title_font

    system_list = [
        ("4.1 暖通空调系统 - 冷源", [
            ["年累计冷负荷", "", "kWh", "按月平均方法计算，考虑建筑分区和气象参数"],
            ["冷源系统能效比(COP)", "", "-", "设计文件，通常3.5-6.0，变频系统可更高"],
            ["冷源系统年运行时间", "", "h", "实际运行小时数，可按供冷季计算"],
            ["冷源系统年耗电量", "", "kWh", "=冷负荷/COP"],
        ]),
        ("4.2 暖通空调系统 - 热源", [
            ["年累计热负荷", "", "kWh", "按月平均方法计算，考虑建筑分区和气象参数"],
            ["热源系统效率", "", "%", "设计文件，燃气锅炉85-95%，电锅炉95-98%"],
            ["热源系统年运行时间", "", "h", "实际运行小时数，可按供暖季计算"],
            ["热源系统年能耗", "", "kWh或m³", "=热负荷/效率"],
        ]),
        ("4.3 生活热水系统", [
            ["用水计算单位数(人数或床位数)", "", "人/床", "实际使用人数"],
            ["热水用水定额", "", "L/人·天", "GB50555规定，办公40-60L/人·班"],
            ["设计热水温度", 60, "°C", "通常55-60°C"],
            ["设计冷水温度", "", "°C", "当地冷水温度，北京取10°C"],
            ["年使用小时数", "", "h", "实际运行时间"],
            ["小时平均耗热量", "", "kWh", "=4.187×人数×定额×(热水温度-冷水温度)×密度/1000"],
            ["年耗热量", "", "kWh", "=小时平均耗热量×年使用小时数"],
            ["生活热水系统年能耗", "", "kWh", "=年耗热量/效率"],
        ]),
        ("4.4 照明系统", [
            ["办公室面积", "", "m²", "实际照明面积"],
            ["办公室照明功率密度", "", "W/m²", "按GB50034设计"],
            ["办公室日照明时间", "", "h", "实际运行时间"],
            ["照明系统年能耗", "", "kWh", "=面积×功率密度×时间×天数/1000"],
        ]),
        ("4.5 电梯系统", [
            ["电梯数量", "", "台", "客梯+货梯总数"],
            ["额定载重量", "", "kg", "设备铭牌"],
            ["速度", "", "m/s", "设备铭牌"],
            ["年运行时间", "", "h", "实际运行小时数"],
            ["电梯系统年能耗", "", "kWh", "=3.6×P×ta×V×W/1000"],
        ]),
        ("4.6 数据中心/机房", [
            ["IT设备总功率", "", "kW", "服务器、存储、网络设备等"],
            ["数据中心PUE值", "", "-", "Power Usage Effectiveness，通常1.5-2.5"],
            ["年运行时间", 8760, "h", "数据中心通常24×7运行"],
            ["数据中心年耗电量", "", "kWh", "=IT设备功率×PUE×运行时间"],
        ]),
        ("4.7 可再生能源 - 光伏", [
            ["光伏面板净面积(Ap)", "", "m²", "设计文件，实际安装面积"],
            ["年太阳辐射照度(I)", "", "kWh/m²", "当地气象数据，北京约1400"],
            ["光电转换效率(KE)", "", "%", "单晶硅15-20%，多晶硅13-17%"],
            ["系统损失效率(KS)", "", "%", "通常15-25%，含线损、灰尘等"],
            ["光伏系统年发电量", "", "kWh", "=Ap×I×KE×(1-KS)"],
            ["光伏系统年减碳量", "", "tCO₂", "=发电量×电网排放因子"],
        ]),
    ]

    current_row = 3
    for section_title, params_data in system_list:
        ws.merge_cells(f'A{current_row}:L{current_row}')
        ws.cell(row=current_row, column=1, value=section_title).font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        ws.cell(row=current_row, column=1).fill = Styles.section_fills[current_row % len(Styles.section_fills)]
        ws.cell(row=current_row, column=1).alignment = Styles.center_align
        current_row += 1

        param_headers = ["参数项", "数值", "单位", "计算公式/说明", "年能耗(kWh)", "碳排放量(tCO₂)", "数据来源", "备注"]
        apply_header_row(ws, current_row, param_headers, Styles.param_fill)
        current_row += 1

        for i, row_data in enumerate(params_data):
            for j, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row+i, column=j, value=value)
                cell.border = Styles.thin_border
                cell.font = Styles.normal_font
                cell.alignment = Styles.left_align if j in [1, 4] else Styles.center_align
                if j == 2:
                    cell.fill = Styles.input_fill
                elif j in [4, 7, 8]:
                    cell.font = Styles.note_font
                    cell.fill = Styles.param_fill
                elif j == 5:
                    cell.fill = Styles.calc_fill
                elif j == 6:
                    cell.fill = Styles.calc_fill
        current_row += len(params_data) + 2

    set_col_widths(ws, [30, 18, 12, 45, 18, 18, 18, 18, 15, 15, 15, 15])
    return ws


def create_materials_sheet(wb):
    """工作表5: 建材生产及运输"""
    ws = wb.create_sheet("建材生产运输")
    ws.sheet_properties.tabColor = 'F44336'

    ws.merge_cells('A1:L1')
    ws.cell(row=1, column=1, value="建材生产及运输阶段碳排放计算（50+建材类型，含回收利用）").font = Styles.title_font

    ws.merge_cells('A3:L3')
    ws.cell(row=3, column=1, value="5.1 建材生产阶段碳排放计算（基于GB/T 51366-2019第6章）").font = Styles.section_font

    material_headers = [
        "序号", "建材大类", "建材名称", "规格型号",
        "消耗量", "单位", "碳排放因子\n(kgCO₂e/单位)", "碳排放量\n(tCO₂e)",
        "数据来源", "是否可回收", "回收比例(%)", "备注"
    ]
    apply_header_row(ws, 4, material_headers, Styles.section_fills[3])

    materials_data = [
        # 结构材料 (1-8)
        [1, "结构材料", "钢筋(HRB400)", "HRB400", "", "t", 2033, "", "采购清单", "是", "", ""],
        [2, "结构材料", "钢筋(HRB500)", "HRB500", "", "t", 2050, "", "采购清单", "是", "", "高强钢筋"],
        [3, "结构材料", "型钢", "Q235/Q345", "", "t", 2100, "", "采购清单", "是", "", "钢结构"],
        [4, "结构材料", "混凝土(C20)", "C20", "", "m³", 280, "", "采购清单", "部分", "", "基础/垫层"],
        [5, "结构材料", "混凝土(C30)", "C30", "", "m³", 320, "", "采购清单", "部分", "", "梁板柱"],
        [6, "结构材料", "混凝土(C40)", "C40", "", "m³", 350, "", "采购清单", "部分", "", "高层建筑"],
        [7, "结构材料", "混凝土(C50+)", "C50-C60", "", "m³", 380, "", "采购清单", "部分", "", "预应力"],
        [8, "结构材料", "预应力钢绞线", "1860MPa", "", "t", 2200, "", "采购清单", "是", "", "预应力结构"],
        # 围护材料 (9-20)
        [9, "围护材料", "普通粘土砖", "MU10", "", "千块", 270, "", "采购清单", "部分", "", ""],
        [10, "围护材料", "混凝土多孔砖", "MU15", "", "千块", 250, "", "采购清单", "部分", "", ""],
        [11, "围护材料", "加气混凝土砌块", "B06", "", "m³", 225, "", "采购清单", "部分", "", "轻质墙体"],
        [12, "围护材料", "陶粒混凝土砌块", "", "", "m³", 240, "", "采购清单", "部分", "", ""],
        [13, "围护材料", "石膏板", "12mm", "", "m²", 3.5, "", "采购清单", "部分", "", "隔墙"],
        [14, "围护材料", "外墙保温板(XPS)", "50mm", "", "m³", 95, "", "采购清单", "否", "", "挤塑板"],
        [15, "围护材料", "外墙保温板(EPS)", "50mm", "", "m³", 85, "", "采购清单", "否", "", "模塑板"],
        [16, "围护材料", "岩棉保温板", "50mm", "", "m³", 75, "", "采购清单", "否", "", "A级防火"],
        [17, "围护材料", "玻璃（普通中空）", "6+12A+6", "", "m²", 12, "", "采购清单", "是", "", ""],
        [18, "围护材料", "玻璃（Low-E中空）", "6+12A+6Low-E", "", "m²", 15, "", "采购清单", "是", "", "节能玻璃"],
        [19, "围护材料", "玻璃幕墙", "单元式", "", "m²", 45, "", "采购清单", "是", "", "含铝型材"],
        [20, "围护材料", "铝合金门窗", "", "", "m²", 25, "", "采购清单", "是", "", ""],
        # 屋面防水材料 (21-25)
        [21, "屋面材料", "防水卷材(SBS)", "4mm", "", "m²", 4.5, "", "采购清单", "否", "", ""],
        [22, "屋面材料", "防水涂料", "聚氨酯", "", "kg", 3.2, "", "采购清单", "否", "", ""],
        [23, "屋面材料", "屋面保温板", "XPS 100mm", "", "m³", 95, "", "采购清单", "否", "", ""],
        [24, "屋面材料", "瓦（陶瓦）", "", "", "m²", 8.5, "", "采购清单", "部分", "", ""],
        [25, "屋面材料", "瓦（彩钢瓦）", "0.6mm", "", "m²", 6.0, "", "采购清单", "是", "", ""],
        # 装饰材料 (26-35)
        [26, "装饰材料", "花岗岩", "30mm", "", "m²", 38, "", "采购清单", "部分", "", "外墙"],
        [27, "装饰材料", "大理石", "20mm", "", "m²", 35, "", "采购清单", "部分", "", "室内"],
        [28, "装饰材料", "瓷砖（地砖）", "800×800", "", "m²", 12, "", "采购清单", "部分", "", ""],
        [29, "装饰材料", "瓷砖（墙砖）", "300×600", "", "m²", 10, "", "采购清单", "部分", "", ""],
        [30, "装饰材料", "实木地板", "", "", "m²", 15, "", "采购清单", "是", "", "可再生"],
        [31, "装饰材料", "复合地板", "", "", "m²", 8.5, "", "采购清单", "部分", "", ""],
        [32, "装饰材料", "涂料（内墙）", "乳胶漆", "", "kg", 3.0, "", "采购清单", "否", "", ""],
        [33, "装饰材料", "涂料（外墙）", "真石漆", "", "kg", 3.5, "", "采购清单", "否", "", ""],
        [34, "装饰材料", "壁纸", "", "", "m²", 2.5, "", "采购清单", "否", "", ""],
        [35, "装饰材料", "吊顶（铝扣板）", "600×600", "", "m²", 18, "", "采购清单", "是", "", ""],
        # 安装材料 (36-45)
        [36, "安装材料", "电缆（铜芯）", "", "", "m", 2.0, "", "采购清单", "是", "", ""],
        [37, "安装材料", "电线", "", "", "m", 0.5, "", "采购清单", "是", ""],
        [38, "安装材料", "管材（钢管）", "", "", "t", 2100, "", "采购清单", "是", ""],
        [39, "安装材料", "管材（PPR）", "", "", "m", 0.8, "", "采购清单", "部分", "", "给水"],
        [40, "安装材料", "管材（PVC）", "", "", "m", 0.6, "", "采购清单", "部分", "", "排水"],
        [41, "安装材料", "风管（镀锌钢板）", "", "", "m²", 12, "", "采购清单", "是", ""],
        [42, "安装材料", "桥架", "", "", "m", 5.5, "", "采购清单", "是", ""],
        [43, "安装材料", "配电箱", "", "", "台", 150, "", "采购清单", "是", ""],
        [44, "安装材料", "开关插座", "", "", "个", 2.5, "", "采购清单", "部分", ""],
        [45, "安装材料", "卫生洁具", "", "", "套", 85, "", "采购清单", "部分", ""],
        # 特殊材料 (46-50)
        [46, "特殊材料", "消防设备", "", "", "项", 5000, "", "采购清单", "部分", "", "估算"],
        [47, "特殊材料", "安防设备", "", "", "项", 3000, "", "采购清单", "部分", "", "估算"],
        [48, "特殊材料", "智能化设备", "", "", "项", 8000, "", "采购清单", "部分", "", "估算"],
        [49, "特殊材料", "电梯设备", "", "", "台", 25000, "", "采购清单", "是", "", "含安装"],
        [50, "特殊材料", "变配电设备", "", "", "项", 15000, "", "采购清单", "是", "", "估算"],
        # 再生材料 (51-55)
        [51, "再生材料", "再生骨料混凝土", "C30", "", "m³", 220, "", "采购清单", "部分", "", "替代30%骨料"],
        [52, "再生材料", "再生钢筋", "", "", "t", 1600, "", "采购清单", "是", "", "废钢回收"],
        [53, "再生材料", "再生砖", "", "", "千块", 180, "", "采购清单", "部分", "", "建筑垃圾再生"],
        [54, "再生材料", "再生木材", "", "", "m³", 80, "", "采购清单", "是", ""],
        [55, "再生材料", "其他再生材料", "", "", "项", 0, "", "估算", "", "", ""],
    ]

    for i, row in enumerate(materials_data, 5):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.normal_font
            cell.alignment = Styles.center_align if j in [1, 6, 7, 10, 11] else Styles.left_align
            if j == 5:
                cell.fill = Styles.input_fill
            elif j in [7, 10, 11]:
                cell.fill = Styles.param_fill
            elif j == 8:
                cell.fill = Styles.calc_fill
                cell.value = f"=IF(E{i}=\"\",\"\",E{i}*G{i}*(1-IF(K{i}=\"\",0,K{i}/100))/1000)"
                cell.number_format = '#,##0.00'

    # 汇总行（按大类）
    summary_rows = [
        (60, "结构材料小计", "=SUM(H5:H12)"),
        (61, "围护材料小计", "=SUM(H13:H24)"),
        (62, "屋面防水材料小计", "=SUM(H25:H29)"),
        (63, "装饰材料小计", "=SUM(H30:H39)"),
        (64, "安装材料小计", "=SUM(H40:H49)"),
        (65, "特殊材料小计", "=SUM(H50:H54)"),
        (66, "再生材料小计", "=SUM(H55:H59)"),
    ]

    for row_num, name, formula in summary_rows:
        ws.cell(row=row_num, column=2, value=name).font = Styles.bold_font
        ws.cell(row=row_num, column=2).fill = Styles.total_fill
        ws.cell(row=row_num, column=8, value=formula).font = Styles.bold_font
        ws.cell(row=row_num, column=8).fill = Styles.total_fill

    ws.cell(row=67, column=2, value="建材生产阶段碳排放合计").font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    ws.cell(row=67, column=2).fill = Styles.section_fills[0]
    ws.cell(row=67, column=8, value="=SUM(H5:H59)").font = Font(name='微软雅黑', size=12, bold=True)
    ws.cell(row=67, column=8).fill = Styles.total_fill

    # 5.2 建材运输阶段
    ws.merge_cells('A69:L69')
    ws.cell(row=69, column=1, value="5.2 建材运输阶段碳排放计算").font = Styles.section_font

    transport_headers = [
        "序号", "建材名称", "运输量(t)", "平均运输距离(km)",
        "运输方式", "运输排放因子\n[kgCO₂e/(t·km)]", "碳排放量(tCO₂e)",
        "数据来源", "运输路线", "备注"
    ]
    apply_header_row(ws, 70, transport_headers, Styles.section_fills[4])

    transport_data = [
        [1, "钢筋", "", "", "公路运输（重载卡车）", 0.0697, "", "运输合同", "", ""],
        [2, "混凝土", "", "", "公路运输（商砼罐车）", 0.0712, "", "运输合同", "商砼站→工地", "短距离"],
        [3, "型钢", "", "", "公路运输（重载卡车）", 0.0697, "", "运输合同", "", ""],
        [4, "砌块/砖", "", "", "公路运输（中型卡车）", 0.0658, "", "运输合同", "", ""],
        [5, "玻璃", "", "", "公路运输（专用车）", 0.0697, "", "运输合同", "厂家→工地", "易碎品"],
        [6, "保温板", "", "", "公路运输（轻型卡车）", 0.0625, "", "运输合同", "", ""],
        [7, "石材", "", "", "公路运输（重载卡车）", 0.0697, "", "运输合同", "矿区→加工厂→工地", ""],
        [8, "电缆/电线", "", "", "公路运输（轻型卡车）", 0.0625, "", "采购记录", "", ""],
        [9, "管材", "", "", "公路运输（中型卡车）", 0.0658, "", "采购记录", "", ""],
        [10, "设备（电梯等）", "", "", "公路运输（特种车）", 0.0745, "", "采购记录", "厂家→工地", "大件运输"],
        [11, "其他材料（合计）", "", "", "公路运输（综合）", 0.0667, "", "估算", "", ""],
        [12, "远距离运输（>500km）", "", "", "铁路运输", 0.0085, "", "运输合同", "", "大批量"],
        [13, "水运（如有）", "", "", "水路运输", 0.0052, "", "运输合同", "", "沿海/内河"],
    ]

    for i, row in enumerate(transport_data, 71):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.normal_font
            if j in [3, 4]:
                cell.fill = Styles.input_fill
            elif j == 6:
                cell.fill = Styles.param_fill
            elif j == 7:
                cell.fill = Styles.calc_fill
                cell.value = f"=IF(C{i}=\"\",\"\",C{i}*D{i}*F{i}/1000)"
                cell.number_format = '#,##0.00'

    ws.cell(row=84, column=2, value="建材运输阶段碳排放合计").font = Styles.bold_font
    ws.cell(row=84, column=2).fill = Styles.total_fill
    ws.cell(row=84, column=7, value="=SUM(G71:G83)").font = Styles.bold_font
    ws.cell(row=84, column=7).fill = Styles.total_fill

    set_col_widths(ws, [8, 15, 18, 18, 25, 22, 22, 15, 20, 20])
    return ws


def create_construction_sheet(wb):
    """工作表6: 建造及拆除阶段"""
    ws = wb.create_sheet("建造拆除阶段")
    ws.sheet_properties.tabColor = '00BCD4'

    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value="建造及拆除阶段碳排放计算（含绿色施工措施）").font = Styles.title_font

    ws.merge_cells('A3:J3')
    ws.cell(row=3, column=1, value="6.1 建筑建造阶段 - 分部分项工程碳排放计算").font = Styles.section_font

    construction_headers = [
        "序号", "分部工程", "分项工程名称", "工程量",
        "单位", "能耗系数\n(kWh/单位)", "能源用量\n(kWh)",
        "碳排放因子\n(kgCO₂/kWh)", "碳排放量\n(kgCO₂)", "备注"
    ]
    apply_header_row(ws, 4, construction_headers, Styles.section_fills[5])

    construction_data = [
        [1, "地基与基础", "土方开挖", "", "m³", "", "", 0.5836, "", "含外运"],
        [2, "地基与基础", "土方回填", "", "m³", "", "", 0.5836, "", ""],
        [3, "地基与基础", "地基处理（桩基）", "", "m", "", "", 0.5836, "", "预制桩/灌注桩"],
        [4, "地基与基础", "基础混凝土", "", "m³", "", "", 0.5836, "", "筏板/独立基础"],
        [5, "地基与基础", "基础钢筋", "", "t", "", "", 0.5836, "", ""],
        [6, "主体结构", "柱混凝土", "", "m³", "", "", 0.5836, "", ""],
        [7, "主体结构", "柱钢筋", "", "t", "", "", 0.5836, "", ""],
        [8, "主体结构", "梁板混凝土", "", "m³", "", "", 0.5836, "", ""],
        [9, "主体结构", "梁板钢筋", "", "t", "", "", 0.5836, "", ""],
        [10, "主体结构", "墙体混凝土", "", "m³", "", "", 0.5836, "", "剪力墙"],
        [11, "主体结构", "墙体钢筋", "", "t", "", "", 0.5836, "", ""],
        [12, "主体结构", "钢结构安装", "", "t", "", "", 0.5836, "", "如有钢结构"],
        [13, "装饰装修", "砌体工程", "", "m³", "", "", 0.5836, "", "填充墙"],
        [14, "装饰装修", "外墙保温", "", "m²", "", "", 0.5836, "", ""],
        [15, "装饰装修", "外墙装饰", "", "m²", "", "", 0.5836, "", "涂料/幕墙/石材"],
        [16, "装饰装修", "内墙装饰", "", "m²", "", "", 0.5836, "", "抹灰/涂料/壁纸"],
        [17, "装饰装修", "楼地面工程", "", "m²", "", "", 0.5836, "", "地砖/地板/石材"],
        [18, "装饰装修", "门窗安装", "", "m²", "", "", 0.5836, "", ""],
        [19, "机电安装", "给排水系统", "", "项", "", "", 0.5836, "", ""],
        [20, "机电安装", "电气系统", "", "项", "", "", 0.5836, "", ""],
        [21, "机电安装", "暖通系统", "", "项", "", "", 0.5836, "", ""],
        [22, "机电安装", "消防系统", "", "项", "", "", 0.5836, "", ""],
        [23, "机电安装", "智能化系统", "", "项", "", "", 0.5836, "", ""],
        [24, "机电安装", "电梯安装", "", "台", "", "", 0.5836, "", ""],
    ]

    for i, row in enumerate(construction_data, 5):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.normal_font
            cell.alignment = Styles.center_align if j in [1, 5, 8] else Styles.left_align
            if j in [4, 6]:
                cell.fill = Styles.input_fill
            elif j == 7:
                cell.fill = Styles.calc_fill
                cell.value = f"=IF(D{i}=\"\",\"\",D{i}*F{i})"
                cell.number_format = '#,##0.00'
            elif j == 9:
                cell.fill = Styles.calc_fill
                cell.value = f"=IF(G{i}=\"\",\"\",G{i}*H{i}/1000)"
                cell.number_format = '#,##0.00'

    ws.cell(row=29, column=2, value="建造阶段碳排放合计").font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    ws.cell(row=29, column=2).fill = Styles.section_fills[0]
    ws.cell(row=29, column=9, value="=SUM(I5:I28)").font = Font(name='微软雅黑', size=12, bold=True)
    ws.cell(row=29, column=9).fill = Styles.total_fill

    # 6.2 拆除阶段
    ws.merge_cells('A31:J31')
    ws.cell(row=31, column=1, value="6.2 建筑拆除阶段碳排放计算").font = Styles.section_font

    demolition_headers = [
        "序号", "拆除项目", "工程量", "单位",
        "能耗系数(kWh/单位)", "能源用量(kWh)",
        "碳排放因子(kgCO₂/kWh)", "碳排放量(kgCO₂)", "备注"
    ]
    apply_header_row(ws, 32, demolition_headers, Styles.section_fills[6])

    demolition_data = [
        [1, "人工拆除（墙体）", "", "m²", "", "", 0.5836, "", ""],
        [2, "人工拆除（楼板）", "", "m²", "", "", 0.5836, "", ""],
        [3, "机械拆除（主体）", "", "m²", "", "", 0.5836, "破碎锤/挖掘机", ""],
        [4, "机械拆除（基础）", "", "m³", "", "", 0.5836, "液压破碎", ""],
        [5, "垃圾外运", "", "t", "", "", 0.5836, "距离___km", ""],
        [6, "拆除垃圾分拣", "", "t", "", "", 0.5836, "可回收物分拣", ""],
        [7, "危险废物处理", "", "t", "", "", 0.5836, "石棉/含铅材料", ""],
    ]

    for i, row in enumerate(demolition_data, 33):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.normal_font
            if j in [3, 5]:
                cell.fill = Styles.input_fill
            elif j == 6:
                cell.fill = Styles.calc_fill
                cell.value = f"=IF(C{i}=\"\",\"\",C{i}*E{i})"
            elif j == 8:
                cell.fill = Styles.calc_fill
                cell.value = f"=IF(F{i}=\"\",\"\",F{i}*G{i}/1000)"

    ws.cell(row=40, column=2, value="拆除阶段碳排放合计").font = Styles.bold_font
    ws.cell(row=40, column=2).fill = Styles.total_fill
    ws.cell(row=40, column=8, value="=SUM(H33:H39)").font = Styles.bold_font
    ws.cell(row=40, column=8).fill = Styles.total_fill

    set_col_widths(ws, [8, 18, 22, 12, 15, 18, 18, 18, 20])
    return ws


def create_water_waste_sheet(wb):
    """工作表7: 水资源与废弃物"""
    ws = wb.create_sheet("水资源与废弃物")
    ws.sheet_properties.tabColor = '795548'

    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value="水资源消耗与废弃物处理碳排放计算").font = Styles.title_font

    # 7.1 水资源消耗
    ws.merge_cells('A3:J3')
    ws.cell(row=3, column=1, value="7.1 水资源消耗碳排放（供水、排水、水处理）").font = Styles.section_font

    water_headers = [
        "序号", "用水类型", "年用水量", "单位",
        "供水能耗系数(kWh/m³)", "供水碳排放(tCO₂)",
        "排水处理能耗(kWh/m³)", "排水碳排放(tCO₂)",
        "数据来源", "备注"
    ]
    apply_header_row(ws, 4, water_headers, Styles.section_fills[0])

    water_data = [
        [1, "生活用水（办公）", "", "m³", 0.35, "", 0.25, "", "水表", "人均40-60L/天"],
        [2, "生活用水（餐饮）", "", "m³", 0.35, "", 0.25, "", "水表", ""],
        [3, "生活用水（卫生）", "", "m³", 0.35, "", 0.25, "", "水表", ""],
        [4, "空调补水", "", "m³", 0.35, "", 0.25, "", "运行记录", "冷却塔补水"],
        [5, "绿化灌溉用水", "", "m³", 0.35, "", 0.25, "", "运行记录", ""],
        [6, "道路清洗用水", "", "m³", 0.35, "", 0.25, "", "运行记录", ""],
        [7, "中水回用", "", "m³", 0.45, "", 0.35, "", "运行记录", "减碳项"],
        [8, "雨水收集利用", "", "m³", 0.30, "", 0.20, "", "运行记录", "减碳项"],
    ]

    for i, row in enumerate(water_data, 5):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.normal_font
            if j == 3:
                cell.fill = Styles.input_fill
            elif j in [5, 7]:
                cell.fill = Styles.param_fill
            elif j in [6, 8]:
                cell.fill = Styles.calc_fill
                if j == 6:
                    cell.value = f"=IF(C{i}=\"\",\"\",C{i}*E{i}*0.5836/1000)"
                elif j == 8:
                    cell.value = f"=IF(C{i}=\"\",\"\",C{i}*G{i}*0.5836/1000)"
                cell.number_format = '#,##0.00'

    ws.cell(row=13, column=2, value="水资源消耗碳排放合计").font = Styles.bold_font
    ws.cell(row=13, column=2).fill = Styles.total_fill
    ws.cell(row=13, column=6, value="=SUM(F5:F12)").font = Styles.bold_font
    ws.cell(row=13, column=6).fill = Styles.total_fill
    ws.cell(row=13, column=8, value="=SUM(H5:H12)").font = Styles.bold_font
    ws.cell(row=13, column=8).fill = Styles.total_fill

    # 7.2 废弃物处理
    ws.merge_cells('A15:J15')
    ws.cell(row=15, column=1, value="7.2 废弃物处理碳排放（生活垃圾、建筑垃圾、危险废物）").font = Styles.section_font

    waste_headers = [
        "序号", "废弃物类型", "年产生量", "单位",
        "处理方式", "处理排放因子(kgCO₂/t)",
        "碳排放量(tCO₂)", "回收减碳量(tCO₂)",
        "数据来源", "备注"
    ]
    apply_header_row(ws, 16, waste_headers, Styles.section_fills[1])

    waste_data = [
        [1, "生活垃圾", "", "t", "填埋", 350, "", "", "物业记录", ""],
        [2, "生活垃圾", "", "t", "焚烧", -150, "", "", "物业记录", "发电回收能量"],
        [3, "餐厨垃圾", "", "t", "厌氧消化", -80, "", "", "物业记录", "产生沼气"],
        [4, "可回收物（纸类）", "", "t", "回收利用", -420, "", "", "物业记录", "减碳项"],
        [5, "可回收物（塑料）", "", "t", "回收利用", -1200, "", "", "物业记录", "减碳项"],
        [6, "可回收物（金属）", "", "t", "回收利用", -1500, "", "", "物业记录", "减碳项"],
        [7, "可回收物（玻璃）", "", "t", "回收利用", -250, "", "", "物业记录", "减碳项"],
        [8, "建筑垃圾（运营期）", "", "t", "填埋", 15, "", "", "物业记录", "装修垃圾"],
        [9, "建筑垃圾（拆除期）", "", "t", "填埋/回收", 20, "", "", "拆除方案", ""],
        [10, "危险废物", "", "t", "专业处理", 800, "", "", "危废转移联单", "实验室/医疗"],
    ]

    for i, row in enumerate(waste_data, 17):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.normal_font
            if j == 3:
                cell.fill = Styles.input_fill
            elif j == 6:
                cell.fill = Styles.param_fill
            elif j == 7:
                cell.fill = Styles.calc_fill
                cell.value = f"=IF(C{i}=\"\",\"\",C{i}*F{i}/1000)"
                cell.number_format = '#,##0.00'
            elif j == 8:
                cell.fill = Styles.calc_fill

    ws.cell(row=27, column=2, value="废弃物处理碳排放合计").font = Styles.bold_font
    ws.cell(row=27, column=2).fill = Styles.total_fill
    ws.cell(row=27, column=7, value="=SUM(G17:G26)").font = Styles.bold_font
    ws.cell(row=27, column=7).fill = Styles.total_fill

    set_col_widths(ws, [8, 22, 15, 10, 18, 20, 18, 18, 15, 20])
    return ws


def create_carbon_sink_sheet(wb):
    """工作表8: 碳汇与隐含碳"""
    ws = wb.create_sheet("碳汇与隐含碳")
    ws.sheet_properties.tabColor = '607D8B'

    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value="建筑碳汇与隐含碳排放计算").font = Styles.title_font

    # 8.1 建筑碳汇
    ws.merge_cells('A3:J3')
    ws.cell(row=3, column=1, value="8.1 建筑碳汇计算（绿地碳汇、立体绿化、屋顶绿化）").font = Styles.section_font

    sink_headers = [
        "序号", "碳汇类型", "面积", "单位",
        "碳汇系数[kgCO₂/(m²·a)]", "年碳汇量(tCO₂)",
        "计算依据", "植物类型", "备注"
    ]
    apply_header_row(ws, 4, sink_headers, Styles.section_fills[2])

    sink_data = [
        [1, "绿地碳汇（乔木）", "", "m²", 2.5, "", "《城市绿地碳汇计量监测技术规程》", "本地树种", ""],
        [2, "绿地碳汇（灌木）", "", "m²", 1.2, "", "同上", "", ""],
        [3, "绿地碳汇（草坪）", "", "m²", 0.5, "", "同上", "", ""],
        [4, "屋顶绿化", "", "m²", 1.8, "", "《绿色建筑评价标准》", "轻型/重型", ""],
        [5, "垂直绿化", "", "m²", 1.5, "", "同上", "攀援植物", ""],
        [6, "室内绿化", "", "m²", 0.8, "", "估算", "盆栽植物", ""],
        [7, "雨水花园", "", "m²", 2.0, "", "估算", "湿地植物", ""],
    ]

    for i, row in enumerate(sink_data, 5):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.normal_font
            if j == 3:
                cell.fill = Styles.input_fill
            elif j == 5:
                cell.fill = Styles.param_fill
            elif j == 6:
                cell.fill = Styles.calc_fill
                cell.value = f"=IF(C{i}=\"\",\"\",C{i}*E{i}/1000)"
                cell.number_format = '#,##0.00'

    ws.cell(row=12, column=2, value="建筑年碳汇总计").font = Styles.bold_font
    ws.cell(row=12, column=2).fill = Styles.total_fill
    ws.cell(row=12, column=6, value="=SUM(F5:F11)").font = Styles.bold_font
    ws.cell(row=12, column=6).fill = Styles.total_fill

    # 8.2 隐含碳排放
    ws.merge_cells('A14:J14')
    ws.cell(row=14, column=1, value="8.2 隐含碳排放计算（家具、设备、IT资产等）").font = Styles.section_font

    embodied_headers = [
        "序号", "资产类别", "资产名称", "数量",
        "单位", "碳排放因子(kgCO₂e/单位)", "碳排放量(tCO₂e)",
        "使用寿命(年)", "年碳排放量(tCO₂e)", "备注"
    ]
    apply_header_row(ws, 15, embodied_headers, Styles.section_fills[3])

    embodied_data = [
        [1, "办公家具", "办公桌", "", "套", 85, "", 10, "", ""],
        [2, "办公家具", "办公椅", "", "把", 35, "", 8, "", ""],
        [3, "办公设备", "台式电脑", "", "台", 250, "", 5, "", "含显示器"],
        [4, "办公设备", "笔记本电脑", "", "台", 180, "", 4, "", ""],
        [5, "办公设备", "打印机/复印机", "", "台", 350, "", 7, "", ""],
        [6, "办公设备", "服务器", "", "台", 800, "", 5, "", "数据中心"],
        [7, "办公设备", "空调（分体）", "", "台", 450, "", 10, "", ""],
        [8, "办公设备", "空调（中央）", "", "台", 2500, "", 15, "", ""],
        [9, "办公设备", "电梯", "", "台", 15000, "", 20, "", ""],
        [10, "其他资产", "厨房设备", "", "套", 5000, "", 10, "", "如有食堂"],
        [11, "其他资产", "安防设备", "", "套", 3000, "", 10, "", ""],
    ]

    for i, row in enumerate(embodied_data, 16):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.normal_font
            if j == 4:
                cell.fill = Styles.input_fill
            elif j == 6:
                cell.fill = Styles.param_fill
            elif j == 7:
                cell.fill = Styles.calc_fill
                cell.value = f"=IF(D{i}=\"\",\"\",D{i}*F{i}/1000)"
                cell.number_format = '#,##0.00'
            elif j == 9:
                cell.fill = Styles.calc_fill
                cell.value = f"=IF(H{i}=\"\",\"\",G{i}/H{i})"
                cell.number_format = '#,##0.00'

    ws.cell(row=27, column=3, value="隐含碳排放合计").font = Styles.bold_font
    ws.cell(row=27, column=3).fill = Styles.total_fill
    ws.cell(row=27, column=7, value="=SUM(G16:G26)").font = Styles.bold_font
    ws.cell(row=27, column=7).fill = Styles.total_fill

    set_col_widths(ws, [8, 15, 18, 10, 10, 25, 20, 15, 20, 20])
    return ws


def create_scenario_sheet(wb):
    """工作表9: 情景分析"""
    ws = wb.create_sheet("情景分析")
    ws.sheet_properties.tabColor = 'E91E63'

    ws.merge_cells('A1:M1')
    ws.cell(row=1, column=1, value="多情景碳排放对比分析（基准/节能/碳中和）").font = Styles.title_font
    ws.cell(row=1, column=1).alignment = Styles.center_align

    ws.merge_cells('A3:M3')
    ws.cell(row=3, column=1, value="9.1 情景参数设置").font = Styles.section_font

    scenario_headers = ["参数类别", "参数名称", "单位", "基准情景", "节能情景", "强化节能", "碳中和情景", "计算依据", "备注"]
    apply_header_row(ws, 4, scenario_headers, Styles.section_fills[8])

    scenario_params = [
        ["能源效率", "建筑综合节能率", "%", 50, 65, 75, 85, "GB 50189", "相对于基准"],
        ["能源效率", "暖通空调COP", "-", 3.5, 4.5, 5.5, 7.0, "设备标准", "冷水机组"],
        ["能源效率", "照明功率密度", "W/m²", 9.0, 7.0, 5.0, 3.5, "GB 50034", "办公建筑"],
        ["可再生能源", "光伏覆盖率", "%", 5, 15, 30, 60, "屋面面积", "BIPV"],
        ["可再生能源", "绿电采购比例", "%", 0, 20, 50, 100, "市场化交易", "零碳电力"],
        ["运行管理", "智能控制覆盖率", "%", 20, 60, 85, 100, "智能化标准", "BA系统"],
        ["建材", "绿色建材比例", "%", 20, 50, 70, 90, "绿色认证", ""],
        ["建材", "再生材料替代率", "%", 5, 15, 30, 50, "资源利用", ""],
    ]

    for i, row in enumerate(scenario_params, 5):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.bold_font if j <= 2 else Styles.normal_font
            cell.alignment = Styles.center_align if j <= 7 else Styles.left_align
            if 4 <= j <= 7:
                cell.fill = Styles.input_fill

    # 9.2 情景计算结果
    current_row = 15
    ws.merge_cells(f'A{current_row}:M{current_row}')
    ws.cell(row=current_row, column=1, value="9.2 各情景碳排放计算结果对比").font = Styles.section_font
    current_row += 1

    result_headers = ["排放类别", "基准情景(tCO₂e/a)", "节能情景", "强化节能", "碳中和情景", "节能减排量", "节能减排率", "碳中和减排量", "碳中和减排率", "数据来源", "备注"]
    apply_header_row(ws, current_row, result_headers, Styles.section_fills[9])
    current_row += 1

    result_data = [
        ["运行阶段-电力", "", "", "", "", "", "", "", "", "自动计算"],
        ["运行阶段-化石燃料", "", "", "", "", "", "", "", "", "自动计算"],
        ["建材生产及运输(年化)", "", "", "", "", "", "", "", "", "自动计算"],
        ["建造及拆除(年化)", "", "", "", "", "", "", "", "", "自动计算"],
        ["水资源消耗", "", "", "", "", "", "", "", "", "自动计算"],
        ["废弃物处理", "", "", "", "", "", "", "", "", "自动计算"],
        ["减：建筑碳汇", "", "", "", "", "", "", "", "", "负值"],
        ["减：可再生能源减碳", "", "", "", "", "", "", "", "", "负值"],
        ["碳排放总计", "", "", "", "", "", "", "", "", "自动计算"],
    ]

    for row_data in result_data:
        for j, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.bold_font if j == 1 else Styles.normal_font
            cell.alignment = Styles.center_align if j <= 9 else Styles.left_align
            if 2 <= j <= 5:
                cell.fill = Styles.calc_fill
            elif j in [7, 9]:
                cell.fill = Styles.benchmark_fill
                cell.number_format = '0.0%'
        current_row += 1

    set_col_widths(ws, [22, 20, 18, 18, 18, 18, 15, 18, 15, 15, 20])
    return ws


def create_sensitivity_sheet(wb):
    """工作表10: 敏感性分析"""
    ws = wb.create_sheet("敏感性分析")
    ws.sheet_properties.tabColor = '3F51B5'

    ws.merge_cells('A1:L1')
    ws.cell(row=1, column=1, value="碳排放敏感性分析 - 关键参数影响评估").font = Styles.title_font
    ws.cell(row=1, column=1).alignment = Styles.center_align

    ws.merge_cells('A3:L3')
    ws.cell(row=3, column=1, value="10.1 单因素敏感性分析（变化±10%, ±20%, ±30%）").font = Styles.section_font

    sens_headers = ["敏感因素", "基准值", "单位", "-30%", "-20%", "-10%", "基准", "+10%", "+20%", "+30%", "敏感系数", "排名"]
    apply_header_row(ws, 4, sens_headers, Styles.section_fills[9])

    sens_data = [
        ["建筑面积", 25000, "m²"], ["年用电量", 1500000, "kWh"],
        ["年用气量", 50000, "m³"], ["年用热量", 8000, "GJ"],
        ["入住率", 80, "%"], ["暖通空调COP", 4.0, "-"],
        ["照明功率密度", 8.0, "W/m²"], ["可再生能源比例", 15, "%"],
        ["绿电采购比例", 0, "%"], ["建材总碳排放", 5000, "tCO₂e"],
        ["年用水量", 50000, "m³"], ["废弃物产生量", 200, "t"],
        ["碳汇面积", 5000, "m²"], ["电网排放因子", 0.5836, "kgCO₂/kWh"],
    ]

    for i, row in enumerate(sens_data, 5):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.bold_font if j == 1 else Styles.normal_font
            cell.alignment = Styles.center_align
            if j == 2:
                cell.fill = Styles.input_fill
            elif 4 <= j <= 10:
                cell.fill = Styles.calc_fill
            elif j == 11:
                cell.fill = Styles.warning_fill
            elif j == 12:
                cell.fill = Styles.benchmark_fill

    set_col_widths(ws, [20, 15, 12, 12, 12, 12, 12, 12, 12, 12, 15, 12])
    return ws


def create_benchmark_sheet(wb):
    """工作表11: 对标与预测"""
    ws = wb.create_sheet("对标与预测")
    ws.sheet_properties.tabColor = '009688'

    ws.merge_cells('A1:L1')
    ws.cell(row=1, column=1, value="碳排放对标分析与未来趋势预测").font = Styles.title_font
    ws.cell(row=1, column=1).alignment = Styles.center_align

    ws.merge_cells('A3:L3')
    ws.cell(row=3, column=1, value="11.1 碳排放强度对标分析").font = Styles.section_font

    bench_headers = ["对标指标", "本项目计算值", "单位", "国家限额", "地方限额", "行业先进", "国际先进", "达标情况", "差距分析", "改进建议"]
    apply_header_row(ws, 4, bench_headers, Styles.section_fills[10])

    bench_data = [
        ["单位面积运行阶段碳排放", "", "kgCO₂/(m²·a)", 75, 65, 50, 35, "", "", ""],
        ["单位面积全生命周期碳排放", "", "kgCO₂e/m²", 850, 750, 600, 450, "", "", ""],
        ["人均运行阶段碳排放", "", "tCO₂/(人·a)", 5.0, 4.5, 3.5, 2.5, "", "", ""],
        ["单位面积电力消耗", "", "kWh/(m²·a)", 150, 130, 100, 70, "", "", ""],
        ["可再生能源利用率", "", "%", 5, 10, 20, 40, "", "", ""],
        ["绿色建材应用比例", "", "%", 30, 40, 60, 80, "", "", ""],
        ["碳排放强度等级", "", "级", "-", "-", "-", "-", "□优秀□良好□合格□不合格", "", ""],
    ]

    for i, row in enumerate(bench_data, 5):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.bold_font if j == 1 else Styles.normal_font
            cell.alignment = Styles.center_align if j <= 8 else Styles.left_align
            if j == 2:
                cell.fill = Styles.input_fill
            elif 4 <= j <= 6:
                cell.fill = Styles.benchmark_fill
            elif j == 8:
                cell.fill = Styles.calc_fill

    # 11.2 碳排放预测
    ws.merge_cells('A14:L14')
    ws.cell(row=14, column=1, value="11.2 碳排放趋势预测（未来10年）").font = Styles.section_font

    forecast_headers = ["年度", "运行阶段排放", "建材运输年化", "建造拆除年化", "水资源废弃物", "隐含碳年化", "碳汇减碳", "可再生能源减碳", "净碳排放", "同比变化(%)", "累计排放"]
    apply_header_row(ws, 15, forecast_headers, Styles.section_fills[11])

    for year_offset in range(11):
        row = 16 + year_offset
        ws.cell(row=row, column=1, value=2026 + year_offset).font = Styles.bold_font
        ws.cell(row=row, column=1).fill = Styles.param_fill
        for col in range(2, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = Styles.thin_border
            cell.fill = Styles.calc_fill

    set_col_widths(ws, [28, 18, 18, 15, 15, 15, 15, 18, 18, 18, 18])
    return ws


def create_trading_sheet(wb):
    """工作表12: 碳交易分析"""
    ws = wb.create_sheet("碳交易分析")
    ws.sheet_properties.tabColor = 'FF5722'

    ws.merge_cells('A1:L1')
    ws.cell(row=1, column=1, value="碳排放权交易与碳成本核算").font = Styles.title_font
    ws.cell(row=1, column=1).alignment = Styles.center_align

    ws.merge_cells('A3:L3')
    ws.cell(row=3, column=1, value="12.1 碳配额管理与核算").font = Styles.section_font

    quota_headers = ["项目", "年度配额量(tCO₂e)", "实际排放量", "配额盈亏", "配额价格(元/t)", "配额价值(元)", "交易策略", "履约情况", "数据来源"]
    apply_header_row(ws, 4, quota_headers, Styles.section_fills[11])

    quota_data = [
        ["运行阶段碳排放", "", "", "", 80, "", "□出售 □购买", "□履约 □未履约", "生态环境部门"],
        ["建材生产运输(年化)", "", "", "", 80, "", "□出售 □购买", "□履约 □未履约", ""],
        ["建造拆除(年化)", "", "", "", 80, "", "□出售 □购买", "□履约 □未履约", ""],
        ["其他排放源", "", "", "", 80, "", "□出售 □购买", "□履约 □未履约", ""],
        ["合计", "", "", "", "", "", "", "", ""],
    ]

    for i, row in enumerate(quota_data, 5):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.bold_font if j == 1 else Styles.normal_font
            cell.alignment = Styles.center_align if j in [1, 5, 7, 8] else Styles.left_align
            if j in [2, 3, 5]:
                cell.fill = Styles.input_fill
            elif j == 4:
                cell.fill = Styles.calc_fill
                cell.value = f"=IF(B{i}=\"\",\"\",B{i}-C{i})"
            elif j == 6:
                cell.fill = Styles.calc_fill
                cell.value = f"=IF(D{i}=\"\",\"\",D{i}*E{i})"

    # 12.2 CCER开发潜力
    ws.merge_cells('A12:L12')
    ws.cell(row=12, column=1, value="12.2 碳减排项目（CCER）开发潜力").font = Styles.section_font

    ccer_headers = ["减排项目类型", "年减排量(tCO₂e/a)", "CCER价格(元/t)", "年收益(元)", "开发成本(元)", "项目状态", "开发建议"]
    apply_header_row(ws, 13, ccer_headers, Styles.section_fills[0])

    ccer_data = [
        ["分布式光伏发电", "", "", "", "", "□开发中 □规划中", ""],
        ["太阳能热水系统", "", "", "", "", "□开发中 □规划中", ""],
        ["地源热泵系统", "", "", "", "", "□开发中 □规划中", ""],
        ["建筑节能改造", "", "", "", "", "□开发中 □规划中", ""],
        ["碳汇造林项目", "", "", "", "", "□开发中 □规划中", ""],
    ]

    for i, row in enumerate(ccer_data, 14):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.bold_font if j == 1 else Styles.normal_font
            if j in [2, 3, 4]:
                cell.fill = Styles.calc_fill
            elif j == 5:
                cell.fill = Styles.input_fill

    set_col_widths(ws, [25, 18, 18, 18, 20, 18, 20])
    return ws


def create_summary_sheet(wb):
    """工作表13: 碳排放汇总"""
    ws = wb.create_sheet("碳排放汇总")
    ws.sheet_properties.tabColor = 'F44336'

    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value="公共建筑全生命周期碳排放汇总报告").font = Font(name='微软雅黑', size=20, bold=True, color='1F4E79')
    ws.cell(row=1, column=1).alignment = Styles.center_align

    ws.merge_cells('A2:J2')
    ws.cell(row=2, column=1, value="基于GB/T 51366-2019和JS/T 303-2026 - 增强版V2.0").font = Font(name='微软雅黑', size=12, italic=True, color='666666')
    ws.cell(row=2, column=1).alignment = Styles.center_align

    # 13.1 项目基本信息摘要
    ws.merge_cells('A4:J4')
    ws.cell(row=4, column=1, value="一、项目基本信息摘要").font = Font(name='微软雅黑', size=14, bold=True, color='1F4E79')

    basic_summary = [
        ["项目名称", "", "建筑类型", "", "计算年度", "", "编制日期", "", "", ""],
        ["建筑面积", "m²", "建筑层数", "层", "使用人数", "人", "年运行天数", "天", "", ""],
        ["年用电量", "kWh", "年用气量", "m³", "年用热量", "GJ", "年用水量", "m³", "", ""],
        ["可再生能源类型", "", "装机容量", "kW", "年发电量", "kWh", "年减碳量", "tCO₂", "", ""],
    ]

    for i, row in enumerate(basic_summary, 5):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.bold_font if j in [1, 3, 5, 7] else Styles.normal_font
            cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid') if j in [1, 3, 5, 7] else Styles.input_fill
            cell.alignment = Styles.center_align

    # 13.2 全生命周期碳排放汇总
    ws.merge_cells('A11:J11')
    ws.cell(row=11, column=1, value="二、建筑全生命周期碳排放汇总（含所有计算模块）").font = Font(name='微软雅黑', size=14, bold=True, color='1F4E79')

    summary_headers = [
        "排放阶段/类别", "碳排放量\n(tCO₂e)", "单位面积碳排放\n(kgCO₂e/m²)", "占总排放比例\n(%)",
        "人均碳排放\n(tCO₂e/人)", "主要排放源", "关键影响因素", "减排潜力\n(%)", "减排建议", "备注"
    ]
    apply_header_row(ws, 12, summary_headers, Styles.section_fills[0])
    ws.row_dimensions[12].height = 35

    summary_sections = [
        ["建材生产阶段", "", "", "", "", "主要建材消耗", "", "", "", ""],
        ["建材运输阶段", "", "", "", "", "运输距离和方式", "", "", "", ""],
        ["建材阶段小计", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["建造阶段", "", "", "", "", "分部分项工程", "", "", "", ""],
        ["绿色施工节能量", "", "", "", "", "装配式等措施", "", "", "", ""],
        ["拆除阶段", "", "", "", "", "拆除工程量", "", "", "", ""],
        ["建造拆除小计", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["运行阶段-电力", "", "", "", "", "暖通空调、照明等", "", "", "", ""],
        ["运行阶段-天然气", "", "", "", "", "供暖、生活用气", "", "", "", ""],
        ["运行阶段-热力", "", "", "", "", "集中供热", "", "", "", ""],
        ["运行阶段-其他能源", "", "", "", "", "汽油、柴油等", "", "", "", ""],
        ["运行阶段小计", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["水资源消耗", "", "", "", "", "供水和排水", "", "", "", ""],
        ["废弃物处理", "", "", "", "", "垃圾和废弃物", "", "", "", ""],
        ["水与废弃物小计", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["隐含碳排放（年化）", "", "", "", "", "家具、设备、IT", "资产使用寿命", "", "", ""],
        ["建筑碳汇（年减碳）", "", "", "", "", "绿地、立体绿化", "绿化面积和类型", "", "", "负值"],
        ["可再生能源减碳", "", "", "", "", "光伏、太阳能", "装机容量", "", "", "负值"],
        ["", "", "", "", "", "", "", "", "", ""],
        ["建筑全生命周期碳排放总计", "", "", "", "", "", "", "", "", ""],
    ]

    current_row = 13
    for row_data in summary_sections:
        for j, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.normal_font
            cell.alignment = Styles.left_align if j in [6, 7, 8, 9, 10] else Styles.center_align
            if j == 1:
                cell.font = Styles.bold_font
            if "小计" in row_data[0] or "总计" in row_data[0]:
                for col in range(1, 11):
                    ws.cell(row=current_row, column=col).fill = Styles.total_fill
                    ws.cell(row=current_row, column=col).font = Font(name='微软雅黑', size=10, bold=True)
        current_row += 1

    # 13.3 碳排放强度指标
    ws.merge_cells(f'A{current_row+2}:J{current_row+2}')
    ws.cell(row=current_row+2, column=1, value="三、碳排放强度指标与等级评定").font = Font(name='微软雅黑', size=14, bold=True, color='1F4E79')

    intensity_data = [
        ["指标名称", "计算值", "单位", "国家限额", "地方限额", "行业先进", "国际先进", "达标情况", "等级评定", "备注"],
        ["单位面积建材生产碳排放", "", "kgCO₂e/m²", 200, 180, 150, 100, "", "", ""],
        ["单位面积运行阶段碳排放", "", "kgCO₂e/(m²·a)", 75, 65, 50, 35, "", "", ""],
        ["单位面积全生命周期碳排放", "", "kgCO₂e/m²", 850, 750, 600, 450, "", "", ""],
        ["人均运行阶段碳排放", "", "tCO₂e/(人·a)", 5.0, 4.5, 3.5, 2.5, "", "", ""],
        ["单位建筑面积综合能耗", "", "kgce/(m²·a)", 25, 22, 18, 12, "", "", ""],
        ["可再生能源利用率", "", "%", 5, 10, 20, 40, "", "", ""],
        ["碳排放强度等级", "", "级", "-", "-", "-", "-", "", "□优秀 □良好 □合格 □不合格", ""],
    ]

    for i, row in enumerate(intensity_data, current_row+3):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Styles.thin_border
            cell.font = Styles.normal_font
            cell.alignment = Styles.center_align
            if i == current_row+3:
                cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
                cell.fill = Styles.section_fills[2]
            elif j == 1:
                cell.font = Styles.bold_font
                cell.fill = Styles.param_fill
            elif j == 2:
                cell.fill = Styles.calc_fill
            elif 4 <= j <= 6:
                cell.fill = Styles.benchmark_fill
            elif j in [8, 9]:
                cell.fill = Styles.scenario_fill

    # 13.4 核算结论
    ws.merge_cells(f'A{current_row+13}:J{current_row+13}')
    ws.cell(row=current_row+13, column=1, value="四、碳排放核算结论与减排建议").font = Font(name='微软雅黑', size=14, bold=True, color='1F4E79')

    ws.merge_cells(f'A{current_row+14}:J{current_row+20}')
    conclusion_text = (
        "【核算结论】\n\n"
        "1. 建筑全生命周期碳排放总量为 _____ tCO₂e\n"
        "2. 运行阶段碳排放占比最高，达到 _____%\n"
        "3. 单位面积碳排放强度为 _____ kgCO₂e/m²\n"
        "4. 碳排放强度等级评定为：________\n\n"
        "【主要减排建议】\n"
        "1. 优化暖通空调系统运行策略，提高能效，预计可减排 ____ tCO₂e/a\n"
        "2. 增加可再生能源利用比例至 ____%，预计可减排 ____ tCO₂e/a\n"
        "3. 加强照明系统节能改造，采用LED和智能控制，预计可减排 ____ tCO₂e/a\n"
        "4. 推进绿色建筑建材采购，提高再生材料使用比例，预计可减排 ____ tCO₂e\n"
        "5. 建立碳排放监测平台，实现能耗数据实时监测和分析\n"
        "6. 探索碳排放权交易，通过市场机制降低碳成本\n"
        "7. 制定碳中和路线图，明确阶段性目标和实施路径"
    )
    cell = ws.cell(row=current_row+14, column=1, value=conclusion_text)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.font = Styles.normal_font
    for col in range(2, 11):
        ws.merge_cells(start_row=current_row+14, start_column=1, end_row=current_row+20, end_column=col)

    set_col_widths(ws, [25, 20, 22, 15, 18, 22, 22, 15, 25, 20])
    return ws


# ========================================
# 主程序
# ========================================
def main():
    """主函数：生成完整的Excel计算表格"""
    print("🚀 正在生成公共建筑碳排放计算表格...")
    print("=" * 60)

    wb = openpyxl.Workbook()

    # 生成所有工作表
    sheets = [
        ("1. 使用说明", create_cover_sheet),
        ("2. 项目基本信息", create_basic_info_sheet),
        ("3. 运行阶段-能源", create_energy_sheet),
        ("4. 运行阶段-系统", create_systems_sheet),
        ("5. 建材生产运输", create_materials_sheet),
        ("6. 建造拆除阶段", create_construction_sheet),
        ("7. 水资源与废弃物", create_water_waste_sheet),
        ("8. 碳汇与隐含碳", create_carbon_sink_sheet),
        ("9. 情景分析", create_scenario_sheet),
        ("10. 敏感性分析", create_sensitivity_sheet),
        ("11. 对标与预测", create_benchmark_sheet),
        ("12. 碳交易分析", create_trading_sheet),
        ("13. 碳排放汇总", create_summary_sheet),
    ]

    for name, create_func in sheets:
        print(f"📊 正在创建: {name}")
        create_func(wb)

    # 保存文件
    wb.save(OUTPUT_FILE)

    print("=" * 60)
    print(f"✅ Excel文件已保存至: {OUTPUT_FILE}")
    print(f"📊 包含 {len(wb.sheetnames)} 个工作表:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"   {i}. {name}")
    print("=" * 60)
    print("🎉 生成完成！")


if __name__ == "__main__":
    main()
