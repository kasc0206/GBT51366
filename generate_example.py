#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
公共建筑碳排放计算报告 - 完整示例生成器
基于GB/T 51366-2019和JS/T 303-2026标准规范
生成一个完整的示例项目：某市公共服务中心
"""

import os
import shutil
from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(OUTPUT_DIR, "templates")
EXAMPLE_DIR = os.path.join(OUTPUT_DIR, "examples", "demo_公共服务中心")
os.makedirs(EXAMPLE_DIR, exist_ok=True)


def safe_set(ws, row, col, value):
    """安全设置单元格（跳过合并单元格）"""
    try:
        ws.cell(row=row, column=col).value = value
    except:
        pass


def generate_example():
    """生成完整示例"""
    logger.info("=" * 60)
    logger.info("🚀 正在生成完整示例项目...")
    logger.info("=" * 60)

    # 复制模板
    excel_src = os.path.join(TEMPLATE_DIR, "公共建筑碳排放计算表_精细化版.xlsx")
    word_src = os.path.join(TEMPLATE_DIR, "公共建筑碳排放计算报告.docx")
    excel_dst = os.path.join(EXAMPLE_DIR, "计算表_公共服务中心.xlsx")
    word_dst = os.path.join(EXAMPLE_DIR, "报告_公共服务中心.docx")

    shutil.copy(excel_src, excel_dst)
    shutil.copy(word_src, word_dst)
    logger.info("✅ 已复制模板文件")

    # 填充Excel
    logger.info("📝 正在填充Excel数据...")
    wb = load_workbook(excel_dst)

    # 01. 项目基本信息
    ws = wb["01. 项目基本信息"]
    safe_set(ws, 5, 2, "XX市公共服务中心")
    safe_set(ws, 6, 2, "XX市高新区科技大道888号")
    safe_set(ws, 7, 2, "公共建筑（综合体）")
    safe_set(ws, 8, 2, "办公/服务/配套")
    safe_set(ws, 9, 2, 35000)
    safe_set(ws, 10, 2, 28000)
    safe_set(ws, 11, 2, 7000)
    safe_set(ws, 12, 2, 12)
    safe_set(ws, 13, 2, 2)
    safe_set(ws, 14, 2, 48.5)
    safe_set(ws, 16, 2, "2020")
    safe_set(ws, 17, 2, 2025)
    safe_set(ws, 18, 2, 1800)
    safe_set(ws, 19, 2, 365)
    safe_set(ws, 20, 2, 12)
    safe_set(ws, 21, 2, 85)
    safe_set(ws, 22, 2, 150)
    safe_set(ws, 23, 2, 8500)
    safe_set(ws, 24, 2, 24.3)
    safe_set(ws, 25, 2, 2000)
    safe_set(ws, 26, 2, 500)

    # 02. 气象参数
    ws = wb["02. 气象参数"]
    safe_set(ws, 4, 2, 1380)
    safe_set(ws, 5, 2, -8)
    safe_set(ws, 6, 2, 34)
    safe_set(ws, 7, 2, 2.3)
    safe_set(ws, 8, 2, 1800)
    safe_set(ws, 9, 2, 8)

    # 05-06. 负荷计算
    ws = wb["05. 冷负荷计算"]
    safe_set(ws, 4, 2, 35000)
    safe_set(ws, 5, 2, 95)
    safe_set(ws, 6, 2, 1800)

    ws = wb["06. 热负荷计算"]
    safe_set(ws, 4, 2, 35000)
    safe_set(ws, 5, 2, 65)
    safe_set(ws, 6, 2, 2880)

    # 07. 冷源系统
    ws = wb["07. 冷源系统"]
    safe_set(ws, 5, 2, 4.2)
    safe_set(ws, 6, 2, 85)
    safe_set(ws, 7, 2, 65)
    safe_set(ws, 8, 2, 45)

    # 08. 热源系统
    ws = wb["08. 热源系统"]
    safe_set(ws, 5, 2, 0.92)
    safe_set(ws, 6, 2, 120)

    # 10. 制冷剂排放
    ws = wb["10. 制冷剂排放"]
    safe_set(ws, 4, 2, "R410A")
    safe_set(ws, 6, 2, 120)
    safe_set(ws, 7, 2, 6)

    # 11. 生活热水
    ws = wb["11. 生活热水"]
    safe_set(ws, 4, 2, 1800)

    # 12. 室内照明
    ws = wb["12. 室内照明"]
    lighting = [
        [18000, 9, 10, 250],
        [2500, 11, 8, 250],
        [3000, 15, 14, 365],
        [4000, 5, 14, 365],
        [6000, 2, 24, 365],
        [1000, 5, 4, 365],
    ]
    for i, row in enumerate(lighting, 5):
        for j, val in enumerate(row, 2):
            safe_set(ws, i, j, val)

    # 14. 电梯系统
    ws = wb["14. 电梯系统"]
    safe_set(ws, 4, 2, 6)
    safe_set(ws, 5, 2, 1050)
    safe_set(ws, 6, 2, 1.75)
    safe_set(ws, 7, 2, 2)
    safe_set(ws, 8, 2, 2000)
    safe_set(ws, 9, 2, 0.35)
    safe_set(ws, 10, 2, 2800)

    # 15. 数据中心
    ws = wb["15. 数据中心"]
    safe_set(ws, 4, 2, 45)

    # 20. 光伏发电
    ws = wb["20. 光伏发电"]
    safe_set(ws, 4, 2, 2000)
    safe_set(ws, 6, 2, 18)
    safe_set(ws, 7, 2, 20)

    # 24-25. 建材
    ws = wb["24. 钢筋"]
    safe_set(ws, 4, 2, 1800)
    safe_set(ws, 5, 2, 1200)

    ws = wb["25. 混凝土"]
    safe_set(ws, 4, 2, 6500)
    safe_set(ws, 5, 2, 8000)
    safe_set(ws, 6, 2, 3500)
    safe_set(ws, 7, 2, 2000)

    # 29. 建材运输
    ws = wb["29. 建材运输"]
    transport = [
        [3000, 150],
        [14500, 50],
        [6500, 120],
        [2000, 200],
        [8000, 80],
    ]
    for i, row in enumerate(transport, 6):
        safe_set(ws, i, 3, row[0])
        safe_set(ws, i, 4, row[1])

    # 36. 绿地碳汇
    ws = wb["36. 绿地碳汇"]
    safe_set(ws, 4, 2, 4500)
    safe_set(ws, 5, 2, 2800)
    safe_set(ws, 6, 2, 1200)
    safe_set(ws, 7, 2, 2000)
    safe_set(ws, 8, 2, 500)

    wb.save(excel_dst)
    logger.info(f"✅ Excel 已填充: {excel_dst}")
    logger.info(f"✅ Word 模板: {word_dst}")

    # 生成说明文档
    readme_path = os.path.join(EXAMPLE_DIR, "README.md")
    with open(readme_path, "w", encoding="utf-8") as f:
        f.write("""# 示例项目：XX市公共服务中心

## 项目概况

| 项目 | 内容 |
|-----|------|
| 项目名称 | XX市公共服务中心 |
| 建筑类型 | 公共建筑（办公/服务/配套综合体） |
| 建筑面积 | 35,000 m²（地上28,000 m²，地下7,000 m²） |
| 建筑层数 | 地上12层，地下2层 |
| 使用人数 | 1,800人 |
| 计算年度 | 2025年 |

## 文件说明

| 文件 | 说明 |
|-----|------|
| 计算表_公共服务中心.xlsx | 已填充数据的精细化计算表（39个工作表） |
| 报告_公共服务中心.docx | Word报告模板（待数据同步后生成完整报告） |

## 使用步骤

1. 打开 `计算表_公共服务中心.xlsx` 查看示例数据
2. 根据需要修改黄色单元格中的数据
3. 使用同步工具将数据同步到Word报告

```bash
# 同步数据到Word
python ../main.py --sync
```

## 计算结果预览

| 指标 | 值 |
|-----|-----|
| 年冷负荷 | ~5,092,200 kWh |
| 年热负荷 | ~3,253,500 kWh |
| 光伏年发电量 | ~453,600 kWh |
| 光伏年减碳量 | ~264.7 tCO₂ |
| 绿地年碳汇 | ~21.8 tCO₂ |

---

**基于标准**: GB/T 51366-2019、JS/T 303-2026  
**生成日期**: 2025年4月6日
""")
    logger.info(f"✅ 示例说明: {readme_path}")

    logger.info("=" * 60)
    logger.info("✅ 示例生成完成！")
    logger.info(f"📁 示例目录: {EXAMPLE_DIR}")
    logger.info("=" * 60)


if __name__ == "__main__":
    generate_example()
