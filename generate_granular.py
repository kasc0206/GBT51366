#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
公共建筑碳排放计算 - 精细化计算单元（35个独立工作表）
将计算拆分为独立的、专注的计算单元，每个单元只负责一个具体的计算任务。

设计理念：
  - 单一职责：每个工作表只计算一个具体的内容
  - 独立完整：每个工作表都是自包含的，有输入、公式、输出
  - 易于扩展：新增计算单元只需添加一个独立的工作表
  - 易于维护：修改某个计算不影响其他计算单元

用法：
  python generate_granular.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ========================================
# 配置
# ========================================
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "公共建筑碳排放计算表_精细化版.xlsx")

# ========================================
# 样式系统
# ========================================
S = type('Styles', (), {
    'title': Font(name='微软雅黑', size=14, bold=True, color='1F4E79'),
    'section': Font(name='微软雅黑', size=12, bold=True, color='FFFFFF'),
    'normal': Font(name='微软雅黑', size=10),
    'bold': Font(name='微软雅黑', size=10, bold=True),
    'note': Font(name='微软雅黑', size=9, italic=True, color='666666'),
    'input': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    'calc': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    'total': PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid'),
    'param': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'),
    'header': PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid'),
    'border': Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    ),
    'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
    'left': Alignment(horizontal='left', vertical='center', wrap_text=True),
})


def create_sheet(wb, name, color, title, width=20):
    """创建标准化的工作表"""
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = color
    
    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1, value=title).font = S.title
    ws.cell(row=1, column=1).alignment = S.center
    
    ws.column_dimensions['A'].width = width
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    return ws


def add_input_section(ws, row, title, headers, data, color_idx=0):
    """添加输入数据区域"""
    colors = ['2E75B6', '4CAF50', 'FF9800', '9C27B0', 'F44336', 
              '00BCD4', '795548', '607D8B', 'E91E63', '3F51B5']
    
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1, value=title).font = S.section
    ws.cell(row=row, column=1).fill = PatternFill(
        start_color=colors[color_idx % len(colors)], 
        end_color=colors[color_idx % len(colors)], 
        fill_type='solid'
    )
    ws.cell(row=row, column=1).alignment = S.center
    row += 1
    
    # 表头
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.font = S.bold
        cell.fill = S.param
        cell.border = S.border
        cell.alignment = S.center
    row += 1
    
    # 数据行
    for d in data:
        for j, v in enumerate(d, 1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.border = S.border
            cell.font = S.normal
            cell.alignment = S.center if j <= 3 else S.left
            if j == 2:
                cell.fill = S.input
            elif j >= 3 and j <= 5:
                cell.fill = S.param
        row += 1
    
    return row


def add_formula_row(ws, row, label, formula, unit="", note=""):
    """添加公式计算行"""
    ws.cell(row=row, column=1, value=label).font = S.bold
    ws.cell(row=row, column=1).border = S.border
    ws.cell(row=row, column=1).fill = S.total
    
    if unit:
        ws.cell(row=row, column=2, value=unit).border = S.border
        ws.cell(row=row, column=2).fill = S.total
        ws.cell(row=row, column=2).alignment = S.center
    
    ws.cell(row=row, column=3, value=formula).border = S.border
    ws.cell(row=row, column=3).fill = S.calc
    ws.cell(row=row, column=3).font = Font(name='微软雅黑', size=10, italic=True, color='0066CC')
    ws.cell(row=row, column=3).alignment = S.left
    
    if note:
        ws.cell(row=row, column=4, value=note).border = S.border
        ws.cell(row=row, column=4).fill = S.param
        ws.cell(row=row, column=4).font = S.note
        ws.cell(row=row, column=4).alignment = S.left
    
    return row


def add_result_row(ws, row, label, formula, unit="", note=""):
    """添加结果行"""
    ws.cell(row=row, column=1, value=label).font = S.bold
    ws.cell(row=row, column=1).border = S.border
    ws.cell(row=row, column=1).fill = S.total
    
    if unit:
        ws.cell(row=row, column=2, value=unit).border = S.border
        ws.cell(row=row, column=2).fill = S.total
        ws.cell(row=row, column=2).alignment = S.center
    
    ws.cell(row=row, column=3, value=formula).border = S.border
    ws.cell(row=row, column=3).fill = S.calc
    ws.cell(row=row, column=3).font = S.bold
    ws.cell(row=row, column=3).alignment = S.center
    
    if note:
        ws.cell(row=row, column=4, value=note).border = S.border
        ws.cell(row=row, column=4).fill = S.param
        ws.cell(row=row, column=4).font = S.note
    
    return row


# ========================================
# 工作表生成器（每个都是独立的计算单元）
# ========================================

def create_cover(wb):
    """00. 封面"""
    ws = wb.active
    ws.title = "00. 封面"
    ws.sheet_properties.tabColor = '1F4E79'
    
    content = [
        ("公共建筑碳排放计算系统 V3.0", 20, True),
        ("精细化计算单元版（35个独立工作表）", 14, False),
        ("", 12, False),
        ("基于 GB/T 51366-2019《建筑碳排放计算标准》", 13, True),
        ("基于 JS/T 303-2026《公共机构碳排放核算指南》", 13, True),
        ("", 12, False),
        ("设计理念", 16, True),
        ("单一职责：每个工作表只计算一个具体内容", 11, False),
        ("独立完整：每个工作表都是自包含的", 11, False),
        ("易于扩展：新增计算单元只需添加独立工作表", 11, False),
        ("易于维护：修改某个计算不影响其他计算单元", 11, False),
    ]
    
    for i, (text, size, bold) in enumerate(content, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(name='微软雅黑', size=size, bold=bold)
        cell.alignment = S.center if i <= 5 else S.left
        ws.merge_cells(f'A{i}:H{i}')
    
    ws.column_dimensions['A'].width = 50
    for c in range(2, 9):
        ws.column_dimensions[get_column_letter(c)].width = 15


def create_01_basic(wb):
    """01. 项目基本信息"""
    ws = create_sheet(wb, "01. 项目基本信息", '4CAF50', 
                     "项目基本信息", 22)
    
    row = 3
    row = add_input_section(ws, row, "基础信息",
        ["项目", "数值", "单位", "数据来源", "说明"],
        [
            ["项目名称", "", "-", "项目文件", "填写完整项目名称"],
            ["建筑地址", "", "-", "规划许可证", "详细到街道门牌号"],
            ["建筑类型", "", "-", "设计文件", "办公/商业/医院等"],
            ["建筑面积(A)", "", "m²", "房产证", "地上+地下总面积"],
            ["地上建筑面积", "", "m²", "设计图纸", ""],
            ["地下建筑面积", "", "m²", "设计图纸", ""],
            ["建筑高度", "", "m", "设计文件", ""],
            ["设计使用寿命", 50, "年", "设计文件", "GB/T 51366规定"],
        ], 0)
    
    row += 1
    row = add_input_section(ws, row, "运行参数",
        ["项目", "数值", "单位", "数据来源", "说明"],
        [
            ["使用人数", "", "人", "物业记录", "日均使用人数"],
            ["年运行天数", 365, "天", "实际运行", ""],
            ["日运行时间", 24, "小时", "实际运行", ""],
            ["入住率", 80, "%", "物业记录", ""],
            ["绿地面积", "", "m²", "规划许可证", "用于碳汇计算"],
            ["光伏装机容量", "", "kWp", "设计文件", "如有光伏"],
            ["地源热泵容量", "", "kW", "设计文件", "如有地源热泵"],
        ], 1)


def create_02_weather(wb):
    """02. 气象参数"""
    ws = create_sheet(wb, "02. 气象参数", '00BCD4', 
                     "气象参数", 22)
    
    row = 3
    row = add_input_section(ws, row, "气象数据",
        ["参数", "数值", "单位", "数据来源", "说明"],
        [
            ["年太阳辐射照度", 1400, "kWh/m²", "气象局", "北京地区"],
            ["采暖室外计算温度", -9, "°C", "设计规范", ""],
            ["空调室外计算干球温度", 35, "°C", "设计规范", ""],
            ["冬季室外平均风速", 2.5, "m/s", "气象局", ""],
            ["年生活热水使用小时数", "", "h", "实际运行", ""],
            ["设计冷水温度", 10, "°C", "当地水源", ""],
            ["设计热水温度", 60, "°C", "设计要求", ""],
        ], 2)


def create_03_envelope(wb):
    """03. 围护结构参数"""
    ws = create_sheet(wb, "03. 围护结构参数", '795548', 
                     "建筑围护结构热工参数", 22)
    
    row = 3
    row = add_input_section(ws, row, "围护结构",
        ["部位", "传热系数", "单位", "设计要求", "实际值"],
        [
            ["外墙", 0.5, "W/(m²·K)", 0.6, ""],
            ["屋面", 0.4, "W/(m²·K)", 0.5, ""],
            ["外窗", 2.0, "W/(m²·K)", 2.5, "Low-E中空"],
            ["地面", 0.5, "W/(m²·K)", 0.6, ""],
            ["幕墙", 1.8, "W/(m²·K)", 2.0, ""],
        ], 3)


def create_04_indoor_params(wb):
    """04. 室内设计参数"""
    ws = create_sheet(wb, "04. 室内设计参数", '607D8B', 
                     "室内设计参数", 22)
    
    row = 3
    row = add_input_section(ws, row, "设计参数",
        ["参数", "数值", "单位", "标准依据", "说明"],
        [
            ["夏季室内设计温度", 26, "°C", "GB 50189", ""],
            ["冬季室内设计温度", 20, "°C", "GB 50189", ""],
            ["夏季室内相对湿度", 65, "%", "GB 50189", ""],
            ["冬季室内相对湿度", 50, "%", "GB 50189", ""],
            ["人均新风量", 30, "m³/(h·人)", "GB 50189", "办公建筑"],
            ["办公室照明标准", 500, "lux", "GB 50034", ""],
        ], 4)


def create_05_cooling_load(wb):
    """05. 冷负荷计算"""
    ws = create_sheet(wb, "05. 冷负荷计算", '2E75B6', 
                     "冷负荷计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "冷负荷参数",
        ["参数", "数值", "单位", "计算公式", "说明"],
        [
            ["建筑面积", "", "m²", "来自01表", ""],
            ["冷负荷指标", "", "W/m²", "设计文件", "通常80-120"],
            ["年供冷小时数", 1800, "h", "实际运行", ""],
            ["同时使用系数", 0.8, "-", "经验值", ""],
        ], 5)
    
    row += 1
    add_formula_row(ws, row, "年冷负荷",
                   "=B3*B4*B5/1000", "kWh", 
                   "=建筑面积×冷负荷指标×小时数×系数/1000")
    row += 1
    add_formula_row(ws, row, "冷负荷峰值",
                   "=B3*B4/1000", "kW",
                   "=建筑面积×冷负荷指标/1000")


def create_06_heating_load(wb):
    """06. 热负荷计算"""
    ws = create_sheet(wb, "06. 热负荷计算", '4CAF50', 
                     "热负荷计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "热负荷参数",
        ["参数", "数值", "单位", "计算公式", "说明"],
        [
            ["建筑面积", "", "m²", "来自01表", ""],
            ["热负荷指标", "", "W/m²", "设计文件", "通常50-80"],
            ["年供暖小时数", 2880, "h", "实际运行", "120天"],
            ["同时使用系数", 0.75, "-", "经验值", ""],
        ], 6)
    
    row += 1
    add_formula_row(ws, row, "年热负荷",
                   "=B3*B4*B5/1000", "kWh",
                   "=建筑面积×热负荷指标×小时数×系数/1000")


def create_07_cooling_system(wb):
    """07. 冷源系统"""
    ws = create_sheet(wb, "07. 冷源系统", '9C27B0', 
                     "冷源系统能耗计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "冷源参数",
        ["参数", "数值", "单位", "数据来源", "说明"],
        [
            ["年冷负荷", "", "kWh", "来自05表", ""],
            ["冷源类型", "", "-", "设计文件", "螺杆机/离心机"],
            ["能效比(COP)", 4.0, "-", "设备铭牌", "通常3.5-6.0"],
            ["冷冻水泵功率", "", "kW", "设备铭牌", ""],
            ["冷却水泵功率", "", "kW", "设备铭牌", ""],
            ["冷却塔功率", "", "kW", "设备铭牌", ""],
        ], 7)
    
    row += 1
    add_formula_row(ws, row, "冷源耗电量",
                   "=B3/B5", "kWh",
                   "=冷负荷/COP")
    row += 1
    add_formula_row(ws, row, "输配系统耗电量",
                   "=(B6+B7+B8)*B5*24", "kWh",
                   "=水泵和冷却塔功率×运行时间")
    row += 1
    add_result_row(ws, row, "冷源系统总能耗",
                   "=B10+B11", "kWh",
                   "=冷源+输配")


def create_08_heating_system(wb):
    """08. 热源系统"""
    ws = create_sheet(wb, "08. 热源系统", 'F44336', 
                     "热源系统能耗计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "热源参数",
        ["参数", "数值", "单位", "数据来源", "说明"],
        [
            ["年热负荷", "", "kWh", "来自06表", ""],
            ["热源类型", "", "-", "设计文件", "燃气锅炉/热泵"],
            ["热源效率", 0.9, "-", "设备铭牌", "85-95%"],
            ["循环泵功率", "", "kW", "设备铭牌", ""],
        ], 0)
    
    row += 1
    add_formula_row(ws, row, "热源耗能量",
                   "=B3/B5", "kWh或m³",
                   "=热负荷/效率")


def create_09_pumping_system(wb):
    """09. 输配系统"""
    ws = create_sheet(wb, "09. 输配系统", 'FF9800', 
                     "输配系统能耗计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "输配参数",
        ["参数", "数值", "单位", "数据来源", "说明"],
        [
            ["冷冻水泵功率", "", "kW", "设备铭牌", ""],
            ["冷却水泵功率", "", "kW", "设备铭牌", ""],
            ["循环泵功率", "", "kW", "设备铭牌", ""],
            ["年运行时间", 2880, "h", "实际运行", ""],
            ["变频系数", 0.8, "-", "经验值", "0.6-0.9"],
        ], 1)
    
    row += 1
    add_formula_row(ws, row, "输配系统年能耗",
                   "=(B3+B4+B5)*B6*B7", "kWh",
                   "=总功率×运行时间×变频系数")


def create_10_refrigerant(wb):
    """10. 制冷剂排放"""
    ws = create_sheet(wb, "10. 制冷剂排放", 'E91E63', 
                     "制冷剂温室气体排放", 22)
    
    row = 3
    row = add_input_section(ws, row, "制冷剂参数",
        ["参数", "数值", "单位", "数据来源", "说明"],
        [
            ["制冷剂类型", "", "-", "设备铭牌", "R410A/R134a等"],
            ["GWP值", 2088, "-", "IPCC AR5", "R410A=2088"],
            ["充注量", "", "kg/台", "设备铭牌", ""],
            ["设备台数", 1, "台", "设计文件", ""],
            ["设备寿命", 15, "年", "设计文件", ""],
            ["年泄漏率", 5, "%", "经验值", "3-10%"],
        ], 2)
    
    row += 1
    add_formula_row(ws, row, "制冷剂年排放量",
                   "=B4*B5/B7*(B8/100)", "tCO₂e",
                   "=充注量×台数/寿命×泄漏率")


def create_11_hot_water(wb):
    """11. 生活热水"""
    ws = create_sheet(wb, "11. 生活热水", '3F51B5', 
                     "生活热水系统能耗计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "热水参数",
        ["参数", "数值", "单位", "标准/来源", "说明"],
        [
            ["用水人数", "", "人", "物业记录", ""],
            ["用水定额", 50, "L/人·天", "GB 50555", "办公建筑"],
            ["设计热水温度", 60, "°C", "设计要求", ""],
            ["设计冷水温度", 10, "°C", "来自02表", ""],
            ["年使用天数", 365, "天", "实际运行", ""],
            ["热源效率", 0.9, "-", "设备铭牌", ""],
        ], 3)
    
    row += 1
    add_formula_row(ws, row, "小时平均耗热量",
                   "=4.187*B3*B4*(B5-B6)/24", "kWh",
                   "=4.187×人数×定额×温差/24")
    row += 1
    add_formula_row(ws, row, "年耗热量",
                   "=B10*B7", "kWh",
                   "=小时耗热量×天数")
    row += 1
    add_result_row(ws, row, "生活热水年能耗",
                   "=B11/B8", "kWh",
                   "=年耗热量/效率")


def create_12_indoor_lighting(wb):
    """12. 室内照明"""
    ws = create_sheet(wb, "12. 室内照明", '009688', 
                     "室内照明能耗计算", 25)
    
    row = 3
    row = add_input_section(ws, row, "各区域照明",
        ["区域", "面积(m²)", "功率密度(W/m²)", "日运行时间(h)", "年运行天数"],
        [
            ["办公室", "", 9, 10, 250],
            ["会议室", "", 11, 8, 250],
            ["大堂门厅", "", 15, 14, 365],
            ["走廊", "", 5, 14, 365],
            ["地下车库", "", 2, 24, 365],
            ["设备用房", "", 5, 4, 365],
        ], 4)
    
    row += 1
    add_formula_row(ws, row, "照明年能耗",
                   "=Σ(面积×功率密度×时间×天数)/1000", "kWh",
                   "各区域之和")


def create_13_outdoor_lighting(wb):
    """13. 室外照明"""
    ws = create_sheet(wb, "13. 室外照明", 'FFC107', 
                     "室外照明能耗计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "室外照明",
        ["区域", "功率(kW)", "日运行时间(h)", "年运行天数", "说明"],
        [
            ["景观照明", "", 4, 365, ""],
            ["道路照明", "", 10, 365, ""],
            ["停车场照明", "", 12, 365, ""],
            ["应急照明", "", 24, 365, ""],
        ], 5)
    
    row += 1
    add_formula_row(ws, row, "室外照明年能耗",
                   "=Σ(功率×时间×天数)", "kWh",
                   "各区域之和")


def create_14_elevator(wb):
    """14. 电梯系统"""
    ws = create_sheet(wb, "14. 电梯系统", '8BC34A', 
                     "电梯系统能耗计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "电梯参数",
        ["参数", "数值", "单位", "数据来源", "说明"],
        [
            ["客梯数量", "", "台", "设计文件", ""],
            ["客梯载重", "", "kg", "设备铭牌", ""],
            ["客梯速度", "", "m/s", "设备铭牌", ""],
            ["货梯数量", "", "台", "设计文件", ""],
            ["货梯载重", "", "kg", "设备铭牌", ""],
            ["特定能量消耗", "", "mWh/kgm", "设备资料", ""],
            ["年运行时间", "", "h", "实际运行", ""],
            ["待机功率", 50, "W", "设备铭牌", ""],
        ], 6)
    
    row += 1
    add_formula_row(ws, row, "电梯年能耗",
                   "=3.6×P×t×V×W/1000+待机×时间/1000", "kWh",
                   "GB/T 51366-2019公式")


def create_15_datacenter(wb):
    """15. 数据中心"""
    ws = create_sheet(wb, "15. 数据中心", 'CDDC39', 
                     "数据中心能耗计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "数据中心参数",
        ["参数", "数值", "单位", "数据来源", "说明"],
        [
            ["IT设备总功率", "", "kW", "设备清单", ""],
            ["PUE值", 1.8, "-", "设计文件", "绿色<1.5"],
            ["年运行时间", 8760, "h", "实际运行", "24×7"],
        ], 7)
    
    row += 1
    add_formula_row(ws, row, "数据中心年能耗",
                   "=B3*B4*B5", "kWh",
                   "=IT功率×PUE×时间")


def create_16_ev_charging(wb):
    """16. 电动汽车充电桩"""
    ws = create_sheet(wb, "16. 电动汽车充电桩", 'FFEB3B', 
                     "充电桩能耗计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "充电桩参数",
        ["参数", "数值", "单位", "数据来源", "说明"],
        [
            ["慢充桩数量", "", "个", "设计文件", "7kW"],
            ["快充桩数量", "", "个", "设计文件", "60kW"],
            ["慢充使用率", 30, "%", "经验值", ""],
            ["快充使用率", 10, "%", "经验值", ""],
            ["年运行天数", 365, "天", "实际运行", ""],
        ], 0)
    
    row += 1
    add_formula_row(ws, row, "充电桩年能耗",
                   "=慢充×7×使用率×天数+快充×60×使用率×天数", "kWh", "")


def create_17_kitchen(wb):
    """17. 厨房餐饮系统"""
    ws = create_sheet(wb, "17. 厨房餐饮系统", 'FF5722', 
                     "厨房设备能耗计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "厨房设备",
        ["设备", "功率(kW)", "日运行时间(h)", "使用率(%)", "说明"],
        [
            ["灶具", "", 6, 50, ""],
            ["蒸箱", "", 4, 40, ""],
            ["冰箱", "", 24, 100, ""],
            ["排风系统", "", 8, 50, ""],
        ], 1)
    
    row += 1
    add_formula_row(ws, row, "厨房年能耗",
                   "=Σ(功率×时间×使用率×天数)", "kWh", "")


def create_18_water_supply(wb):
    """18. 供水系统"""
    ws = create_sheet(wb, "18. 供水系统", '00BCD4', 
                     "供水系统能耗计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "供水参数",
        ["参数", "数值", "单位", "数据来源", "说明"],
        [
            ["年用水量", "", "m³", "水表", ""],
            ["供水泵功率", "", "kW", "设备铭牌", ""],
            ["供水能耗系数", 0.35, "kWh/m³", "经验值", ""],
        ], 2)
    
    row += 1
    add_formula_row(ws, row, "供水系统年能耗",
                   "=B3*B5", "kWh",
                   "=用水量×能耗系数")


def create_19_ventilation(wb):
    """19. 通风系统"""
    ws = create_sheet(wb, "19. 通风系统", '795548', 
                     "通风系统能耗计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "通风参数",
        ["参数", "数值", "单位", "数据来源", "说明"],
        [
            ["风机总功率", "", "kW", "设备铭牌", ""],
            ["年运行时间", "", "h", "实际运行", ""],
            ["变频系数", 0.75, "-", "经验值", ""],
        ], 3)
    
    row += 1
    add_formula_row(ws, row, "通风系统年能耗",
                   "=B3*B4*B5", "kWh",
                   "=功率×时间×变频系数")


def create_20_pv(wb):
    """20. 光伏发电"""
    ws = create_sheet(wb, "20. 光伏发电", 'FF9800', 
                     "光伏系统发电计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "光伏参数",
        ["参数", "数值", "单位", "数据来源", "说明"],
        [
            ["光伏面板面积", "", "m²", "设计文件", ""],
            ["年辐射照度", 1400, "kWh/m²", "来自02表", ""],
            ["转换效率", 18, "%", "设备铭牌", ""],
            ["系统损失率", 20, "%", "设计文件", ""],
            ["电网排放因子", 0.5836, "kgCO₂/kWh", "区域电网", ""],
        ], 4)
    
    row += 1
    add_formula_row(ws, row, "年发电量",
                   "=B3*B4*(B5/100)*(1-B6/100)", "kWh",
                   "=面积×辐射×效率×(1-损失)")
    row += 1
    add_result_row(ws, row, "年减碳量",
                   "=B10*B7/1000", "tCO₂",
                   "=发电量×排放因子/1000")


def create_21_solar_hotwater(wb):
    """21. 太阳能热水"""
    ws = create_sheet(wb, "21. 太阳能热水", 'F44336', 
                     "太阳能热水系统计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "太阳能热水参数",
        ["参数", "数值", "单位", "数据来源", "说明"],
        [
            ["集热器面积", "", "m²", "设计文件", ""],
            ["年辐照量", "", "MJ/m²", "气象局", ""],
            ["集热效率", 50, "%", "设备铭牌", ""],
            ["热损失率", 10, "%", "设计文件", ""],
        ], 5)
    
    row += 1
    add_formula_row(ws, row, "年供热量",
                   "=B3*B4*(1-B6/100)*(B5/100)/3.6", "kWh",
                   "=面积×辐照×(1-损失)×效率/3.6")


def create_22_ground_source(wb):
    """22. 地源热泵"""
    ws = create_sheet(wb, "22. 地源热泵", '4CAF50', 
                     "地源热泵系统计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "地源热泵参数",
        ["参数", "数值", "单位", "数据来源", "说明"],
        [
            ["装机容量", "", "kW", "设计文件", ""],
            ["制冷COP", 4.5, "-", "设备铭牌", ""],
            ["制热COP", 3.8, "-", "设备铭牌", ""],
            ["年供冷小时", 1800, "h", "实际运行", ""],
            ["年供暖小时", 1200, "h", "实际运行", ""],
        ], 6)
    
    row += 1
    add_formula_row(ws, row, "地源热泵年能耗",
                   "=B3*(B6/B4+B7/B5)", "kWh",
                   "=容量×(冷负荷/COP冷+热负荷/COP热)")


def create_23_energy_summary(wb):
    """23. 能源消耗汇总"""
    ws = create_sheet(wb, "23. 能源消耗汇总", '1F4E79', 
                     "所有能源消耗汇总", 25)
    
    row = 3
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1, value="各系统能源消耗汇总").font = S.section
    ws.cell(row=row, column=1).fill = S.header
    row += 1
    
    headers = ["系统", "能耗量", "单位", "来源工作表", "碳排放因子", "碳排放量", "占比", "备注"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.font = S.bold
        cell.fill = S.param
        cell.border = S.border
        cell.alignment = S.center
    row += 1
    
    systems = [
        ["冷源系统", "", "kWh", "07表", 0.5836, "", "", "kgCO₂/MWh"],
        ["热源系统", "", "kWh", "08表", 0.5836, "", "", ""],
        ["输配系统", "", "kWh", "09表", 0.5836, "", "", ""],
        ["生活热水", "", "kWh", "11表", 0.5836, "", "", ""],
        ["室内照明", "", "kWh", "12表", 0.5836, "", "", ""],
        ["室外照明", "", "kWh", "13表", 0.5836, "", "", ""],
        ["电梯系统", "", "kWh", "14表", 0.5836, "", "", ""],
        ["数据中心", "", "kWh", "15表", 0.5836, "", "", ""],
        ["充电桩", "", "kWh", "16表", 0.5836, "", "", ""],
        ["厨房餐饮", "", "kWh", "17表", 0.5836, "", "", ""],
        ["供水系统", "", "kWh", "18表", 0.5836, "", "", ""],
        ["通风系统", "", "kWh", "19表", 0.5836, "", "", ""],
    ]
    
    for s in systems:
        for j, v in enumerate(s, 1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.border = S.border
            cell.font = S.normal
            cell.alignment = S.center if j <= 3 else S.left
            if j == 2:
                cell.fill = S.input
            elif j == 6:
                cell.fill = S.calc
                cell.value = f"=IF(B{row}=\"\",\"\",B{row}*E{row}/1000)"
                cell.number_format = '#,##0.00'
            elif j == 7:
                cell.fill = S.calc
        row += 1
    
    # 汇总行
    ws.cell(row=row, column=1, value="电力消耗合计").font = S.bold
    ws.cell(row=row, column=1).fill = S.total
    ws.cell(row=row, column=6, value=f"=SUM(F{row-len(systems)}:F{row-1})").font = S.bold
    ws.cell(row=row, column=6).fill = S.total


def create_24_steel(wb):
    """24. 钢筋"""
    ws = create_sheet(wb, "24. 钢筋", '2E75B6', 
                     "钢筋材料碳排放计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "钢筋参数",
        ["规格", "消耗量(t)", "单位", "排放因子(tCO₂e/t)", "回收率(%)", "净碳排放"],
        [
            ["HRB400", "", "t", 2.033, 90, ""],
            ["HRB500", "", "t", 2.050, 90, ""],
            ["型钢", "", "t", 2.100, 95, ""],
        ], 0)
    
    row += 1
    add_formula_row(ws, row, "钢筋总碳排放",
                   "=Σ[消耗量×排放因子×(1-回收率)]", "tCO₂e",
                   "考虑回收利用")


def create_25_concrete(wb):
    """25. 混凝土"""
    ws = create_sheet(wb, "25. 混凝土", '4CAF50', 
                     "混凝土材料碳排放计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "混凝土参数",
        ["标号", "消耗量(m³)", "单位", "排放因子(tCO₂e/m³)", "说明"],
        [
            ["C20", "", "m³", 0.280, "基础/垫层"],
            ["C30", "", "m³", 0.320, "梁板柱"],
            ["C40", "", "m³", 0.350, "高层建筑"],
            ["C50", "", "m³", 0.380, "预应力"],
        ], 1)
    
    row += 1
    add_formula_row(ws, row, "混凝土总碳排放",
                   "=Σ(消耗量×排放因子)", "tCO₂e", "")


def create_26_envelope_material(wb):
    """26. 围护建材"""
    ws = create_sheet(wb, "26. 围护建材", 'FF9800', 
                     "围护材料碳排放计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "围护材料",
        ["材料", "消耗量", "单位", "排放因子", "回收率(%)", "净碳排放"],
        [
            ["普通粘土砖", "", "千块", "0.270 tCO₂e/千块", 30, ""],
            ["加气混凝土砌块", "", "m³", "0.225 tCO₂e/m³", 30, ""],
            ["XPS保温板", "", "m³", "0.095 tCO₂e/m³", 0, ""],
            ["岩棉保温板", "", "m³", "0.075 tCO₂e/m³", 0, ""],
            ["玻璃", "", "m²", "0.015 tCO₂e/m²", 80, ""],
            ["铝合金门窗", "", "m²", "0.025 tCO₂e/m²", 90, ""],
        ], 2)
    
    row += 1
    add_formula_row(ws, row, "围护材料总碳排放",
                   "=Σ[消耗量×排放因子×(1-回收率)]", "tCO₂e", "")


def create_27_decoration_material(wb):
    """27. 装饰材料"""
    ws = create_sheet(wb, "27. 装饰材料", '9C27B0', 
                     "装饰材料碳排放计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "装饰材料",
        ["材料", "消耗量", "单位", "排放因子", "回收率(%)", "净碳排放"],
        [
            ["花岗岩", "", "m²", "0.038 tCO₂e/m²", 50, ""],
            ["大理石", "", "m²", "0.035 tCO₂e/m²", 50, ""],
            ["瓷砖", "", "m²", "0.012 tCO₂e/m²", 30, ""],
            ["实木地板", "", "m²", "0.015 tCO₂e/m²", 80, ""],
            ["涂料", "", "kg", "0.003 tCO₂e/kg", 0, ""],
            ["壁纸", "", "m²", "0.0025 tCO₂e/m²", 0, ""],
            ["吊顶铝扣板", "", "m²", "0.018 tCO₂e/m²", 90, ""],
        ], 3)
    
    row += 1
    add_formula_row(ws, row, "装饰材料总碳排放",
                   "=Σ[消耗量×排放因子×(1-回收率)]", "tCO₂e", "")


def create_28_installation_material(wb):
    """28. 安装材料"""
    ws = create_sheet(wb, "28. 安装材料", 'F44336', 
                     "安装材料碳排放计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "安装材料",
        ["材料", "消耗量", "单位", "排放因子", "回收率(%)", "净碳排放"],
        [
            ["电缆", "", "m", "0.002 tCO₂e/m", 95, ""],
            ["钢管", "", "t", "2.100 tCO₂e/t", 95, ""],
            ["PPR管材", "", "m", "0.0008 tCO₂e/m", 50, ""],
            ["风管", "", "m²", "0.012 tCO₂e/m²", 95, ""],
            ["配电箱", "", "台", "0.150 tCO₂e/台", 90, ""],
        ], 4)
    
    row += 1
    add_formula_row(ws, row, "安装材料总碳排放",
                   "=Σ[消耗量×排放因子×(1-回收率)]", "tCO₂e", "")


def create_29_material_transport(wb):
    """29. 建材运输"""
    ws = create_sheet(wb, "29. 建材运输", '00BCD4', 
                     "建材运输碳排放计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "运输参数",
        ["建材", "运输量(t)", "距离(km)", "运输方式", "排放因子", "碳排放量"],
        [
            ["钢筋", "", "", "公路重载", 0.0697, ""],
            ["混凝土", "", "", "商砼罐车", 0.0712, ""],
            ["砌块", "", "", "公路中载", 0.0658, ""],
            ["玻璃", "", "", "公路专用", 0.0697, ""],
            ["石材", "", "", "公路重载", 0.0697, ""],
        ], 5)
    
    row += 1
    add_formula_row(ws, row, "建材运输总碳排放",
                   "=Σ(运输量×距离×排放因子)", "tCO₂e",
                   "=量×距×因子")


def create_30_earthwork(wb):
    """30. 土方工程"""
    ws = create_sheet(wb, "30. 土方工程", '795548', 
                     "土方工程碳排放计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "土方参数",
        ["项目", "工程量", "单位", "能耗系数(kWh/单位)", "碳排放(kgCO₂)"],
        [
            ["土方开挖", "", "m³", 5.5, ""],
            ["土方回填", "", "m³", 3.2, ""],
            ["土方外运", "", "m³", 4.8, ""],
        ], 6)
    
    row += 1
    add_formula_row(ws, row, "土方工程碳排放",
                   "=Σ(工程量×能耗系数×排放因子)", "tCO₂e",
                   "排放因子=0.5836 kgCO₂/kWh")


def create_31_structure(wb):
    """31. 结构工程"""
    ws = create_sheet(wb, "31. 结构工程", '607D8B', 
                     "结构工程碳排放计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "结构工程",
        ["项目", "工程量", "单位", "能耗系数(kWh/单位)", "碳排放(kgCO₂)"],
        [
            ["基础混凝土", "", "m³", 35, ""],
            ["柱梁板混凝土", "", "m³", 40, ""],
            ["墙体混凝土", "", "m³", 38, ""],
            ["钢筋制作安装", "", "t", 120, ""],
            ["钢结构安装", "", "t", 150, ""],
        ], 7)
    
    row += 1
    add_formula_row(ws, row, "结构工程碳排放",
                   "=Σ(工程量×能耗系数×排放因子)", "tCO₂e", "")


def create_32_decoration(wb):
    """32. 装饰装修"""
    ws = create_sheet(wb, "32. 装饰装修", 'E91E63', 
                     "装饰装修工程碳排放计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "装饰工程",
        ["项目", "工程量", "单位", "能耗系数(kWh/单位)", "碳排放(kgCO₂)"],
        [
            ["砌体工程", "", "m³", 15, ""],
            ["外墙保温", "", "m²", 2.5, ""],
            ["内墙装饰", "", "m²", 1.8, ""],
            ["楼地面工程", "", "m²", 2.2, ""],
            ["门窗安装", "", "m²", 1.5, ""],
        ], 0)
    
    row += 1
    add_formula_row(ws, row, "装饰装修碳排放",
                   "=Σ(工程量×能耗系数×排放因子)", "tCO₂e", "")


def create_33_mep_install(wb):
    """33. 机电安装"""
    ws = create_sheet(wb, "33. 机电安装", '3F51B5', 
                     "机电安装工程碳排放计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "机电安装",
        ["项目", "工程量", "单位", "能耗系数(kWh/单位)", "碳排放(kgCO₂)"],
        [
            ["给排水系统", "", "项", 5000, ""],
            ["电气系统", "", "项", 8000, ""],
            ["暖通系统", "", "项", 12000, ""],
            ["消防系统", "", "项", 3000, ""],
            ["智能化系统", "", "项", 4000, ""],
        ], 1)
    
    row += 1
    add_formula_row(ws, row, "机电安装碳排放",
                   "=Σ(工程量×能耗系数×排放因子)", "tCO₂e", "")


def create_34_demolition(wb):
    """34. 拆除工程"""
    ws = create_sheet(wb, "34. 拆除工程", 'FF5722', 
                     "拆除工程碳排放计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "拆除参数",
        ["项目", "工程量", "单位", "能耗系数(kWh/单位)", "碳排放(kgCO₂)"],
        [
            ["人工拆除墙体", "", "m²", 8, ""],
            ["机械拆除主体", "", "m²", 25, ""],
            ["垃圾外运", "", "t", 15, ""],
            ["垃圾分拣", "", "t", 5, ""],
        ], 2)
    
    row += 1
    add_formula_row(ws, row, "拆除工程碳排放",
                   "=Σ(工程量×能耗系数×排放因子)", "tCO₂e", "")


def create_35_water_waste(wb):
    """35. 水资源与废弃物"""
    ws = create_sheet(wb, "35. 水资源与废弃物", '009688', 
                     "水资源与废弃物碳排放计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "水资源",
        ["用水类型", "年用水量(m³)", "供水能耗(kWh/m³)", "碳排放(tCO₂)", "备注"],
        [
            ["生活用水", "", 0.35, "", ""],
            ["空调补水", "", 0.35, "", ""],
            ["绿化灌溉", "", 0.35, "", ""],
        ], 3)
    
    row += 2
    row = add_input_section(ws, row, "废弃物",
        ["废弃物类型", "年产生量(t)", "处理排放因子(kgCO₂/t)", "碳排放(tCO₂)", "备注"],
        [
            ["生活垃圾填埋", "", 350, "", ""],
            ["餐厨垃圾", "", -80, "", "负值=减碳"],
            ["可回收物", "", -800, "", "负值=减碳"],
            ["建筑垃圾", "", 15, "", ""],
        ], 4)
    
    row += 1
    add_formula_row(ws, row, "水资源与废弃物总碳排放",
                   "=水碳排放+废弃物碳排放", "tCO₂e", "")


def create_36_carbon_sink(wb):
    """36. 绿地碳汇"""
    ws = create_sheet(wb, "36. 绿地碳汇", '8BC34A', 
                     "绿地碳汇计算", 22)
    
    row = 3
    row = add_input_section(ws, row, "碳汇参数",
        ["碳汇类型", "面积(m²)", "碳汇系数[kgCO₂/(m²·a)]", "年碳汇量(tCO₂)", "说明"],
        [
            ["乔木绿地", "", 2.5, "", ""],
            ["灌木绿地", "", 1.2, "", ""],
            ["草坪", "", 0.5, "", ""],
            ["屋顶绿化", "", 1.8, "", ""],
            ["垂直绿化", "", 1.5, "", ""],
        ], 5)
    
    row += 1
    add_formula_row(ws, row, "建筑年碳汇总计",
                   "=Σ(面积×碳汇系数)", "tCO₂",
                   "负值（减碳项）")


def create_37_embodied_carbon(wb):
    """37. 隐含碳排放"""
    ws = create_sheet(wb, "37. 隐含碳排放", 'FFC107', 
                     "家具设备隐含碳排放", 22)
    
    row = 3
    row = add_input_section(ws, row, "设备家具",
        ["资产类别", "数量", "单位", "碳排放因子", "使用寿命(年)", "年碳排放"],
        [
            ["办公桌", "", "套", "85 kgCO₂e/套", 10, ""],
            ["电脑", "", "台", "250 kgCO₂e/台", 5, ""],
            ["打印机", "", "台", "350 kgCO₂e/台", 7, ""],
            ["空调", "", "台", "450 kgCO₂e/台", 10, ""],
        ], 6)
    
    row += 1
    add_formula_row(ws, row, "隐含碳年化排放",
                   "=Σ(数量×排放因子/使用寿命)", "tCO₂e/a", "")


def create_38_final_summary(wb):
    """38. 碳排放汇总"""
    ws = create_sheet(wb, "38. 碳排放汇总", 'F44336', 
                     "全生命周期碳排放汇总", 25)
    
    row = 3
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1, value="全生命周期碳排放汇总").font = S.section
    ws.cell(row=row, column=1).fill = S.header
    row += 1
    
    headers = ["排放阶段", "碳排放量(tCO₂e)", "单位面积(kgCO₂e/m²)", "占比(%)", "来源工作表", "备注"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.font = S.bold
        cell.fill = S.param
        cell.border = S.border
        cell.alignment = S.center
    row += 1
    
    items = [
        ["建材生产阶段", "", "", "", "24-28表", ""],
        ["建材运输阶段", "", "", "", "29表", ""],
        ["土方工程", "", "", "", "30表", ""],
        ["结构工程", "", "", "", "31表", ""],
        ["装饰装修", "", "", "", "32表", ""],
        ["机电安装", "", "", "", "33表", ""],
        ["拆除工程", "", "", "", "34表", ""],
        ["运行阶段", "", "", "", "23表", ""],
        ["水资源与废弃物", "", "", "", "35表", ""],
        ["隐含碳", "", "", "", "37表", ""],
        ["减：绿地碳汇", "", "", "", "36表", "负值"],
        ["总计", "", "", "", "", ""],
    ]
    
    for item in items:
        for j, v in enumerate(item, 1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.border = S.border
            cell.font = S.normal
            cell.alignment = S.left if j in [1, 5, 6] else S.center
            if "总计" in str(v) or "减：" in str(v):
                cell.fill = S.total
                cell.font = S.bold
        row += 1
    
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1, value="碳排放强度指标").font = S.section
    ws.cell(row=row, column=1).fill = S.header
    row += 1
    
    indicators = [
        ["单位面积建材生产碳排放", "", "kgCO₂e/m²", "<200", "", ""],
        ["单位面积运行阶段碳排放", "", "kgCO₂e/(m²·a)", "<75", "", ""],
        ["单位面积全生命周期碳排放", "", "kgCO₂e/m²", "<850", "", ""],
        ["人均运行阶段碳排放", "", "tCO₂/(人·a)", "<5.0", "", ""],
        ["碳排放强度等级", "", "级", "□优秀 □良好 □合格 □不合格", "", ""],
    ]
    
    for ind in indicators:
        for j, v in enumerate(ind, 1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.border = S.border
            cell.font = S.bold if j == 1 else S.normal
            cell.fill = S.param if j == 1 else S.calc


# ========================================
# 主程序
# ========================================
def main():
    print("🚀 正在生成精细化计算单元Excel...")
    print("=" * 60)

    wb = openpyxl.Workbook()

    sheets = [
        ("封面", create_cover),
        ("01. 项目基本信息", create_01_basic),
        ("02. 气象参数", create_02_weather),
        ("03. 围护结构参数", create_03_envelope),
        ("04. 室内设计参数", create_04_indoor_params),
        ("05. 冷负荷计算", create_05_cooling_load),
        ("06. 热负荷计算", create_06_heating_load),
        ("07. 冷源系统", create_07_cooling_system),
        ("08. 热源系统", create_08_heating_system),
        ("09. 输配系统", create_09_pumping_system),
        ("10. 制冷剂排放", create_10_refrigerant),
        ("11. 生活热水", create_11_hot_water),
        ("12. 室内照明", create_12_indoor_lighting),
        ("13. 室外照明", create_13_outdoor_lighting),
        ("14. 电梯系统", create_14_elevator),
        ("15. 数据中心", create_15_datacenter),
        ("16. 电动汽车充电桩", create_16_ev_charging),
        ("17. 厨房餐饮系统", create_17_kitchen),
        ("18. 供水系统", create_18_water_supply),
        ("19. 通风系统", create_19_ventilation),
        ("20. 光伏发电", create_20_pv),
        ("21. 太阳能热水", create_21_solar_hotwater),
        ("22. 地源热泵", create_22_ground_source),
        ("23. 能源消耗汇总", create_23_energy_summary),
        ("24. 钢筋", create_24_steel),
        ("25. 混凝土", create_25_concrete),
        ("26. 围护建材", create_26_envelope_material),
        ("27. 装饰材料", create_27_decoration_material),
        ("28. 安装材料", create_28_installation_material),
        ("29. 建材运输", create_29_material_transport),
        ("30. 土方工程", create_30_earthwork),
        ("31. 结构工程", create_31_structure),
        ("32. 装饰装修", create_32_decoration),
        ("33. 机电安装", create_33_mep_install),
        ("34. 拆除工程", create_34_demolition),
        ("35. 水资源与废弃物", create_35_water_waste),
        ("36. 绿地碳汇", create_36_carbon_sink),
        ("37. 隐含碳排放", create_37_embodied_carbon),
        ("38. 碳排放汇总", create_38_final_summary),
    ]

    for name, create_func in sheets:
        print(f"📊 正在创建: {name}")
        create_func(wb)

    wb.save(OUTPUT_FILE)

    print("=" * 60)
    print(f"✅ 精细化Excel已保存至: {OUTPUT_FILE}")
    print(f"📊 包含 {len(wb.sheetnames)} 个独立计算单元:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"   {i}. {name}")
    print("=" * 60)
    print("🎉 精细化计算单元生成完成！")


if __name__ == "__main__":
    main()
